"""Ticket Details lookup → Excel workbook (one sheet per vendor-type group)."""
from __future__ import annotations

import io
import re

_HEAD_FILL = "1E3A5F"
_BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")


def _cell(v):
    if v is None:
        return None
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v if x not in (None, ""))
    if isinstance(v, dict):
        return str(v)
    return v


def _sheet_title(label: str, used: set[str]) -> str:
    """Excel sheet titles: ≤31 chars, no []:*?/\\, unique."""
    base = (_BAD_SHEET_CHARS.sub(" ", str(label)).strip() or "Sheet")[:31]
    title, i = base, 2
    while title.lower() in used:
        suffix = f" ({i})"
        title = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


def build_ticket_details_xlsx(groups: list[dict]) -> io.BytesIO:
    """groups: [{label, columns:[{header,field}], rows:[{field: value}]}]."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    default = wb.active
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=_HEAD_FILL)
    used: set[str] = set()
    made = 0

    for g in groups:
        cols = g["columns"]
        ws = wb.create_sheet(_sheet_title(g["label"], used))
        ws.append([c["header"] for c in cols])
        for ci in range(1, len(cols) + 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = head_font
            cell.fill = head_fill
        for row in g["rows"]:
            ws.append([_cell(row.get(c["field"])) for c in cols])
        for ci, c in enumerate(cols, start=1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(len(str(c["header"])) + 2, 12), 40)
        made += 1

    if made:
        wb.remove(default)
    else:
        default.title = "No Results"
        default.append(["No tickets found for the given document/ticket number(s)."])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
