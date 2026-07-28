"""BSP statement XLS (.xlsx) exporters.

Two workbooks, mirroring the two PDFs a user uploads for a BSP group:
  - build_bsp_summary_xlsx  → the uploaded Agent Billing Summary (per-airline lines)
  - build_bsp_detailed_xlsx → the detailed statement (one row per settlement txn)

Both fill real data (not a blank template) so the user can download the same data
they see in the "Summary Uploaded" / "Detailed rows" tabs. The detailed workbook is
built in write-only mode because a detailed statement can carry tens of thousands of
rows.
"""
from __future__ import annotations

import io

_HEAD_FILL = "1E3A5F"   # brand navy, matches the on-screen table header


def _f(v) -> float:
    """Decimal/str/None → float (0.0 on failure) so Excel stores real numbers."""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _s(v) -> str:
    return "" if v is None else str(v)


def _join(v) -> str:
    """A JSONB list (tour / alt_document_numbers) → a comma-joined cell."""
    if not v:
        return ""
    if isinstance(v, (list, tuple)):
        return ", ".join(str(x) for x in v if x not in (None, ""))
    return str(v)


def _period(obj) -> str:
    pf, pt = getattr(obj, "period_from", None), getattr(obj, "period_to", None)
    if pf and pt:
        return f"{pf:%d %b %Y} - {pt:%d %b %Y}"
    return _s(pf or pt)


def xlsx_filename(*parts) -> str:
    """Safe .xlsx download filename from a prefix + identifying parts."""
    raw = "-".join(str(p) for p in parts if p)
    safe = "".join(c for c in raw if c.isalnum() or c in (" ", "-", "_")).strip().replace(" ", "_")
    return f"{safe or 'bsp'}.xlsx"


# ── Summary workbook (small — normal mode, styled) ───────────────────────────

_SUMMARY_HEADERS = [
    "Category", "Airline Code", "IATA", "Airline", "FOP",
    "Issues", "Refunds", "Debit Memos", "Credit Memos",
    "Std Comm", "Sup Comm", "Tax on Comm", "Balance Payable", "Docs",
]
_SUMMARY_MONEY = ("issues", "refunds", "debit_memos", "credit_memos",
                  "std_comm", "sup_comm", "tax_on_comm", "balance_payable")


def build_bsp_summary_xlsx(summary, rows) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary (Uploaded)"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=_HEAD_FILL)
    right = Alignment(horizontal="right")
    bold = Font(bold=True)

    ws.append([f"BSP Summary — {_s(summary.agent_name) or _s(summary.agent_code) or 'Agent'}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([
        f"Agent: {_s(summary.agent_code)} {_s(summary.agent_name)}".strip(),
        None, None,
        f"Reference: {_s(summary.reference)}",
    ])
    ws.append([
        f"Period: {_period(summary)}",
        None, None,
        f"Currency: {_s(summary.currency) or 'INR'}",
    ])
    ws.append([f"Reconciliation: {_s(summary.match_status) or 'pending'}"])
    ws.append([])

    ws.append(_SUMMARY_HEADERS)
    header_row = ws.max_row
    for c in range(1, len(_SUMMARY_HEADERS) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = head_font
        cell.fill = head_fill

    totals = {k: 0.0 for k in _SUMMARY_MONEY}
    total_docs = 0
    for r in rows:
        for k in _SUMMARY_MONEY:
            totals[k] += _f(getattr(r, k, None))
        total_docs += int(getattr(r, "doc_count", 0) or 0)
        ws.append([
            _s(r.category),
            _s(r.airline_code),
            _s(r.airline_iata),
            _s(r.airline_name),
            _s(r.fop),
            _f(r.issues), _f(r.refunds), _f(r.debit_memos), _f(r.credit_memos),
            _f(r.std_comm), _f(r.sup_comm), _f(r.tax_on_comm), _f(r.balance_payable),
            int(getattr(r, "doc_count", 0) or 0),
        ])

    ws.append([
        f"Total ({len(rows)})", "", "", "", "",
        round(totals["issues"], 2), round(totals["refunds"], 2),
        round(totals["debit_memos"], 2), round(totals["credit_memos"], 2),
        round(totals["std_comm"], 2), round(totals["sup_comm"], 2),
        round(totals["tax_on_comm"], 2), round(totals["balance_payable"], 2),
        total_docs,
    ])
    for c in range(1, len(_SUMMARY_HEADERS) + 1):
        ws.cell(row=ws.max_row, column=c).font = bold

    # right-align the numeric columns (Issues … Docs = columns 6..14)
    for col in range(6, len(_SUMMARY_HEADERS) + 1):
        for r_i in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=r_i, column=col).alignment = right

    widths = [16, 12, 8, 28, 8, 14, 14, 14, 14, 12, 12, 13, 15, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


# ── Detailed workbook (can be huge — write-only, streamed) ───────────────────

_DETAIL_HEADERS = [
    "Document #", "Ticket No", "Txn", "Air", "Airline", "Date", "CPUI", "NR", "STAT", "FOP",
    "Txn Amt", "Fare", "Tax", "F&C", "Pen", "Net Sales",
    "Std %", "Std Comm", "Supp %", "Supp Disc", "Tax/Comm", "Balance",
    "Tour", "Alt Docs", "SPDR No", "RTDN", "ESAC", "WAVR",
]


def _tax_buckets(taxes) -> tuple[float, float, float]:
    """(tax, fees&charges, penalty) totals from a row's tax-breakup components,
    matching the on-screen TAX / F&C / PEN columns."""
    tax = fee = pen = 0.0
    for t in taxes or []:
        ct = getattr(t, "component_type", None)
        amt = _f(getattr(t, "amount", None))
        if ct == "TAX":
            tax += amt
        elif ct == "FEE":
            fee += amt
        elif ct == "PENALTY":
            pen += amt
    return tax, fee, pen


def build_bsp_detailed_xlsx(statement, rows, taxes_by_row: dict) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.cell import WriteOnlyCell
    from openpyxl.styles import Font, PatternFill

    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Detailed")

    title = WriteOnlyCell(ws, value=f"BSP Detailed — {_s(statement.airline_name) or _s(statement.airline_code) or 'Statement'}")
    title.font = Font(bold=True, size=14)
    ws.append([title])
    ws.append([f"Airline: {_s(statement.airline_code)} {_s(statement.airline_name)}".strip(),
               None, None, f"Period: {_period(statement)}"])
    ws.append([f"Reference: {_s(statement.statement_name)}", None, None,
               f"Rows: {getattr(statement, 'row_count', None) or len(rows)}"])
    ws.append([])

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=_HEAD_FILL)
    header_cells = []
    for h in _DETAIL_HEADERS:
        c = WriteOnlyCell(ws, value=h)
        c.font = head_font
        c.fill = head_fill
        header_cells.append(c)
    ws.append(header_cells)

    for r in rows:
        taxes = taxes_by_row.get(r.id, [])
        tax, fee, pen = _tax_buckets(taxes)
        if pen == 0.0 and not any(getattr(t, "component_type", None) == "PENALTY" for t in taxes):
            pen = _f(r.penalty_amount)   # fallback to the row's penalty column
        doc = _s(r.document_number) or _s(r.ticket_number)
        ticket_no = doc if r.transaction_type == "TKTT" else ""
        ws.append([
            doc,
            ticket_no,
            _s(r.transaction_type),
            _s(r.airline_accounting_code),
            _s(r.airline_name),
            _s(r.issue_date),
            _s(r.cpui),
            _s(r.nr_code),
            _s(r.stat),
            _s(r.form_of_payment),
            _f(r.transaction_amount),
            _f(r.fare_amount),
            round(tax, 2),
            round(fee, 2),
            round(pen, 2),
            _f(r.net_sales),
            _f(r.standard_commission_rate),
            _f(r.standard_commission_amount),
            _f(r.supplier_discount_rate),
            _f(r.supplier_discount_amount),
            _f(r.tax_on_commission),
            _f(r.balance_payable),
            _join(r.tour),
            _join(r.alt_document_numbers),
            _s(r.spdr_no),
            _s(r.rtdn),
            _s(r.esac),
            _s(r.wavr),
        ])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
