"""XLSX exports of the Master Governance masters.

Each master's list page gives the platform admin a "Export XLS" button so they
can pull the whole master into Excel, review it, and come back to correct it.
Every sheet deliberately reuses that master's *upload template* headers, so a
downloaded file can be edited and fed straight back into the same master's
bulk-upload — with one extra leading id column for reference, which the
importers ignore along with any other column they do not know.

One helper here rather than five near-identical copies of the openpyxl dance in
airlines.py / airports.py / classes.py / suppliers.py / iata_commissions.py.
"""

from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import Iterable, Sequence

from fastapi.responses import StreamingResponse

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Excel truncates anything longer, and openpyxl raises on an over-long title.
_MAX_SHEET_TITLE = 31
# Wide free-text columns (addresses, notes) would otherwise push a column out to
# hundreds of characters and make the sheet unreadable.
_MAX_COL_WIDTH = 60


def cell(value: object) -> object:
    """Coerce one ORM attribute into a value openpyxl can write.

    Booleans become the yes/no the bulk-upload parsers read back, and dates the
    ISO form _parse_date expects, so an exported sheet round-trips as an import.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def master_export_response(
    *,
    sheet_title: str,
    filename: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> StreamingResponse:
    """Build the downloadable .xlsx for one master.

    `rows` must yield values positionally matching `headers`.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:_MAX_SHEET_TITLE]

    ws.append(list(headers))
    for header_cell in ws[1]:
        header_cell.font = Font(bold=True)
    # The header stays visible while the admin scrolls a few thousand rows.
    ws.freeze_panes = "A2"

    widths = [len(str(h)) for h in headers]
    for row in rows:
        values = [cell(v) for v in row]
        ws.append(values)
        for i, value in enumerate(values[: len(widths)]):
            widths[i] = max(widths[i], min(len(str(value)), _MAX_COL_WIDTH))
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width + 2

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
