"""Export a saved income statement (one ticket statement) as PDF or XLSX.

Reuses the per-ticket calculated fields (incentive_breakdown, calculated_incentive,
iata_commission) computed during the Run, plus the saved IncomeSummary aggregates.
"""
import io

from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable

# Keep byte-identical to INCENTIVE_TYPE_KEYS in app/api/v1/tickets.py
INCENTIVE_TYPE_KEYS = [
    "PLB", "Super PLB", "Transaction Fee", "Deposit Incentive (DI)",
    "Marketing Fund", "Ancillary", "Frontend", "Backend", "Cashback",
    "Segment Incentive", "Push Action",
]

_GREY = colors.HexColor("#666666")
_LINE = colors.HexColor("#cccccc")
_HEAD = colors.HexColor("#1e3a5f")
_BLACK = colors.black


def _f(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _num(v) -> str:
    return f"{_f(v):,.2f}"


def _passenger(t) -> str:
    return (getattr(t, "pax_name", None)
            or " ".join(x for x in [getattr(t, "first_name", None), getattr(t, "last_name", None)] if x).strip()
            or "-")


def _airline(t) -> str:
    return getattr(t, "airline_name", None) or getattr(t, "airlines_code", None) or "-"


# ── PDF ──────────────────────────────────────────────────────────────────────
def build_income_statement_pdf(summary, tickets, agency: dict | None = None) -> io.BytesIO:
    agency = agency or {}

    title = ParagraphStyle("title", fontName="Helvetica-Bold", fontSize=16, textColor=_BLACK, leading=20)
    sub   = ParagraphStyle("sub", fontName="Helvetica", fontSize=9, textColor=_GREY, leading=13)
    word  = ParagraphStyle("word", fontName="Helvetica-Bold", fontSize=18, textColor=_BLACK, alignment=TA_RIGHT, leading=22)
    smr   = ParagraphStyle("smr", fontName="Helvetica", fontSize=9, textColor=_GREY, alignment=TA_RIGHT, leading=13)

    period = f"{summary.valid_from:%d %b %Y} - {summary.valid_to:%d %b %Y}"
    header = Table(
        [[
            [Paragraph(agency.get("name") or summary.agency or "Agency", title),
             Paragraph(summary.name or "Income Statement", sub),
             Paragraph(f"{summary.statement_type} · {summary.agency} · {period}", sub)],
            [Paragraph("Income Statement", word),
             Paragraph(f"Tickets: {summary.ticket_count}", smr)],
        ]],
        colWidths=[170 * mm, 90 * mm],
    )
    header.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0)]))

    rows = [["Ticket #", "Passenger", "Airline", "Sector", "Sell Fare", "IATA Comm", "Income", "Status"]]
    for t in tickets:
        rows.append([
            getattr(t, "ticket_number", None) or "-",
            _passenger(t),
            _airline(t),
            getattr(t, "sector", None) or "-",
            _num(getattr(t, "sell_fare", None)),
            _num(getattr(t, "iata_commission", None)),
            _num(getattr(t, "calculated_incentive", None)),
            (getattr(t, "ticket_status", None) or "-").capitalize(),
        ])
    rows.append([
        "", "", "", f"Total ({summary.ticket_count})", "",
        _num(summary.iata_commission_total), _num(summary.total_income), "",
    ])

    table = Table(rows, repeatRows=1,
                  colWidths=[26 * mm, 44 * mm, 34 * mm, 22 * mm, 28 * mm, 28 * mm, 28 * mm, 22 * mm])
    last = len(rows) - 1
    table.setStyle(TableStyle([
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTNAME",      (0, last), (-1, last), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("BACKGROUND",    (0, 0), (-1, 0), _HEAD),
        ("ALIGN",         (4, 0), (6, -1), "RIGHT"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",     (0, 1), (-1, -2), 0.25, _LINE),
        ("LINEABOVE",     (0, last), (-1, last), 0.75, _BLACK),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 3),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
    ]))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    doc.build([header, Spacer(1, 5 * mm), HRFlowable(width="100%", thickness=0.5, color=_LINE), Spacer(1, 5 * mm), table])
    buf.seek(0)
    return buf


# ── XLSX ─────────────────────────────────────────────────────────────────────
def build_income_statement_xlsx(summary, tickets) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Income Statement"

    bold = Font(bold=True)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1E3A5F")
    right = Alignment(horizontal="right")

    ws.append([summary.name or "Income Statement"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"{summary.statement_type} · {summary.agency} · "
               f"{summary.valid_from:%d %b %Y} - {summary.valid_to:%d %b %Y} · {summary.ticket_count} tickets"])
    ws.append([])

    headers = ["Ticket #", "Passenger", "Airline", "Airline Code", "Sector", "Sell Fare",
               *INCENTIVE_TYPE_KEYS, "IATA Comm", "Total Income", "Status"]
    ws.append(headers)
    header_row = ws.max_row
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.font = head_font
        cell.fill = head_fill

    inc_sum = {k: 0.0 for k in INCENTIVE_TYPE_KEYS}
    for t in tickets:
        bd = getattr(t, "incentive_breakdown", None) or {}
        for k in INCENTIVE_TYPE_KEYS:
            inc_sum[k] += _f(bd.get(k))
        ws.append([
            getattr(t, "ticket_number", None) or "",
            _passenger(t),
            getattr(t, "airline_name", None) or "",
            getattr(t, "airlines_code", None) or "",
            getattr(t, "sector", None) or "",
            _f(getattr(t, "sell_fare", None)),
            *[round(_f(bd.get(k)), 2) for k in INCENTIVE_TYPE_KEYS],
            _f(getattr(t, "iata_commission", None)),
            _f(getattr(t, "calculated_incentive", None)),
            getattr(t, "ticket_status", None) or "",
        ])

    total_row = [
        "", "", "", "", f"Total ({summary.ticket_count})", "",
        *[round(inc_sum[k], 2) for k in INCENTIVE_TYPE_KEYS],
        _f(summary.iata_commission_total), _f(summary.total_income), "",
    ]
    ws.append(total_row)
    for c in range(1, len(total_row) + 1):
        ws.cell(row=ws.max_row, column=c).font = bold

    # number columns: Sell Fare (6) … Total Income
    for col in range(6, 6 + len(INCENTIVE_TYPE_KEYS) + 2 + 1):
        for r in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=r, column=col).alignment = right

    widths = [16, 24, 22, 12, 12, 14] + [12] * len(INCENTIVE_TYPE_KEYS) + [12, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio
