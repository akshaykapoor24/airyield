"""Dashboard — supplier income accrued on flown revenue (the PLB accrual board).

Two surfaces:

  /dashboard/overview   the executive KPI + chart layer
  /dashboard/accrual/*  the board itself: grid, filters, cell overrides, freeze, export

The engine is services/plb_accrual.py; this module is transport, filtering and
serialisation only.

WHAT WAS REMOVED AND WHY. `/summary`, `/pending-actions` and `/supplier-comparison`
were built on the old uploaded-ticket flow: they materialised every UploadedTicket
in Python, summed `comm_sell + calculated_incentive` per USER, and reported deal
"suppliers" from `Deal.source_agent`, which is the import mechanism, not a vendor.
The signal worth keeping out of them — how much is waiting on somebody — survives
as the `actions` block on /overview, each count linking to the module that owns it.

`/income-filters` and `/income-summary` are UNCHANGED and stay: the Income Summary
page is a separate screen and is not part of this work.
"""
from __future__ import annotations

import io
from collections import defaultdict
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from pydantic import BaseModel, Field
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.dependencies import get_current_user
from app.models.airline import Airline
from app.models.airline_class_master import AirlineClassMaster
from app.models.approval_workflow import ApprovalActionStatus, DealApproval
from app.models.bsp_statement import BspStatement, BspStatementRow
from app.models.deal import Deal, DealDirection, DealStatusType
from app.models.plb_accrual import PlbAccrualInput, PlbAccrualSnapshot, PlbAirlineSetting
from app.models.uploaded_ticket import UploadedTicket
from app.models.user import User
from app.services import plb_accrual as pa

router = APIRouter()


# ── Shared helpers (kept from the previous module — used by income-summary) ───

def _ym(date_str: str | None) -> str | None:
    """Extract YYYY-MM label from a date string like '2025-01-15', '15-01-2025', etc."""
    if not date_str:
        return None
    s = str(date_str).strip()
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    try:
        parts = s.replace("/", "-").split("-")
        if len(parts) == 3 and len(parts[2]) == 4:
            return f"{parts[2]}-{parts[1].zfill(2)}"
    except Exception:
        pass
    return None


def _f(v) -> float:
    """Coerce Decimal / None to float — DB returns NUMERIC as Decimal."""
    return float(v) if v is not None else 0.0


# ══════════════════════════════════════════════════════════════════════════════
# PLB accrual board
# ══════════════════════════════════════════════════════════════════════════════

class PeriodRead(BaseModel):
    key: str
    label: str
    from_: date = Field(alias="from")
    to: date

    model_config = {"populate_by_name": True}


class MonthCellRead(BaseModel):
    ym: str
    flown: float
    source: str          # derived | manual | pooled | none
    in_period: bool
    confirmed: bool
    deflator_pct: float
    commissionable: float
    pool: float


class AccrualRowRead(BaseModel):
    key: str
    deal_id: int | None
    deal_no: str | None
    airline_name: str
    channel: str
    entity: str | None
    lob: str | None
    plb_period_from: date | None
    plb_period_to: date | None
    plb_period_label: str
    flown_confirmed_through: date | None
    basis_label: str
    deflator_pct: float
    deflator_source: str
    plb_rate_pct: float
    plb_rate_source: str
    plb_rate_explain: str
    months: dict[str, MonthCellRead]
    flown_total: float
    flown_in_period: float
    commissionable_base: float
    accrual: float
    accrual_at_risk: float
    status: str
    status_flags: list[str]
    reasons: list[str]


class FrozenRead(BaseModel):
    period_key: str
    frozen_at: datetime
    total_accrual: float | None
    row_count: int
    note: str | None


class AccrualBoardRead(BaseModel):
    period: PeriodRead
    months: list[str]
    basis: str
    rows: list[AccrualRowRead]
    totals: dict
    data_quality: dict
    frozen: FrozenRead | None


def _row_out(r: pa.BoardRow, months: list[str]) -> AccrualRowRead:
    """Serialise a board row, trimming the month map to the visible window.

    The engine keeps a wider month span than the window because slab achievement
    is measured over the whole incentive period; the client only ever renders the
    window, and shipping two years of cells per row would dwarf the payload.
    """
    return AccrualRowRead(
        key=r.key, deal_id=r.deal_id, deal_no=r.deal_no,
        airline_name=r.airline_name, channel=r.channel, entity=r.entity, lob=r.lob,
        plb_period_from=r.plb_period_from, plb_period_to=r.plb_period_to,
        plb_period_label=r.plb_period_label,
        flown_confirmed_through=r.flown_confirmed_through,
        basis_label=r.basis_label,
        deflator_pct=r.deflator_pct, deflator_source=r.deflator_source,
        plb_rate_pct=r.plb_rate_pct, plb_rate_source=r.plb_rate_source,
        plb_rate_explain=r.plb_rate_explain,
        months={
            ym: MonthCellRead(**vars(r.months[ym])) for ym in months if ym in r.months
        },
        flown_total=r.flown_total, flown_in_period=r.flown_in_period,
        commissionable_base=r.commissionable_base, accrual=r.accrual,
        accrual_at_risk=r.accrual_at_risk,
        status=r.status, status_flags=r.status_flags, reasons=r.reasons,
    )


async def _frozen_for(
    db: AsyncSession, user: User, period_key: str,
) -> PlbAccrualSnapshot | None:
    return (await db.execute(
        select(PlbAccrualSnapshot).where(
            PlbAccrualSnapshot.tenant_id == user.tenant_id,
            PlbAccrualSnapshot.created_by_id == user.id,
            PlbAccrualSnapshot.period_key == period_key,
        )
    )).scalar_one_or_none()


async def _board(
    db: AsyncSession,
    user: User,
    period: str | None,
    date_from: date | None,
    date_to: date | None,
    basis: str,
) -> tuple[pa.Period, dict]:
    p = pa.resolve_period(period, date_from, date_to)
    board = await pa.build_board(db, user.tenant_id, user.id, p, basis)
    return p, board


@router.get("/accrual", response_model=AccrualBoardRead)
async def get_accrual_board(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    period: str | None = Query(default=None, description="AMJ-26 | JFM-26 | 2026-04 | FY26-27"),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    basis: str = Query(default="auto", pattern="^(auto|travel|issue)$"),
    airline: list[str] | None = Query(default=None),
    entity: list[str] | None = Query(default=None),
    channel: list[str] | None = Query(default=None),
    lob: list[str] | None = Query(default=None),
    status_flag: list[str] | None = Query(default=None, alias="status"),
    search: str | None = Query(default=None),
):
    """The accrual grid.

    A frozen period is served from its snapshot verbatim — the whole point of
    freezing is that a late statement upload must not restate a booked accrual.
    Filters still apply to a frozen board, so the totals row is recomputed from
    the snapshot's own rows rather than trusting its stored totals.
    """
    p = pa.resolve_period(period, date_from, date_to)
    snap = await _frozen_for(db, current_user, p.key)

    if snap is not None:
        rows = [pa.BoardRow(**{**r, "months": {
            k: pa.MonthCell(**v) for k, v in r["months"].items()
        }}) for r in snap.grid["rows"]]
        months = snap.grid["months"]
        board_basis = snap.grid.get("basis", basis)
        data_quality = snap.grid.get("data_quality", {})
    else:
        board = await pa.build_board(db, current_user.tenant_id, current_user.id, p, basis)
        rows, months = board["rows"], board["months"]
        board_basis, data_quality = board["basis"], board["data_quality"]

    rows = pa.filter_rows(rows, airline, entity, channel, lob, status_flag, search)

    return AccrualBoardRead(
        period=PeriodRead(key=p.key, label=p.label, **{"from": p.start}, to=p.end),
        months=months,
        basis=board_basis,
        rows=[_row_out(r, months) for r in rows],
        totals=pa.compute_totals(rows, months),
        data_quality=data_quality,
        frozen=None if snap is None else FrozenRead(
            period_key=snap.period_key, frozen_at=snap.frozen_at,
            total_accrual=_f(snap.total_accrual), row_count=snap.row_count, note=snap.note,
        ),
    )


class AccrualFiltersResponse(BaseModel):
    airlines: list[str]
    entities: list[str]
    channels: list[str]
    lobs: list[str]
    statuses: list[str]
    periods: list[dict]


@router.get("/accrual/filters", response_model=AccrualFiltersResponse)
async def get_accrual_filters(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    period: str | None = Query(default=None),
):
    """Facets for the filter bar, taken from the deal lines themselves rather than
    the airline master — an airline with no PLB deal is not a useful filter here."""
    p = pa.resolve_period(period)
    lines = await pa.load_deal_lines(db, current_user.tenant_id, current_user.id, p.start, p.end)

    today = date.today()
    periods = []
    for back in range(0, 6):
        m = today.month - back * 3
        y = today.year
        while m <= 0:
            m += 12
            y -= 1
        q = pa.QUARTER_OF_MONTH[m]
        key = f"{q}-{y % 100:02d}"
        if not any(x["key"] == key for x in periods):
            periods.append({"key": key, "label": f"{q} {y % 100:02d}"})

    return AccrualFiltersResponse(
        airlines=sorted({l.airline_name for l in lines if l.airline_name}),
        entities=sorted({l.entity for l in lines if l.entity}),
        channels=sorted({l.channel for l in lines if l.channel}),
        lobs=sorted({l.lob for l in lines if l.lob}),
        statuses=list(pa.STATUS_ORDER),
        periods=periods,
    )


# ── Cell overrides ───────────────────────────────────────────────────────────

class AccrualInputPatch(BaseModel):
    airline_name: str
    entity: str | None = None
    channel: str | None = None
    lob: str | None = None
    ym: str = Field(pattern=r"^\d{4}-\d{2}$")
    # `None` is not "leave alone" — it is "clear this override and go back to the
    # derived value". A field the caller does not want to touch is simply absent,
    # which is why these are read through `model_fields_set`.
    deflator_pct: float | None = None
    plb_rate_pct: float | None = None
    manual_flown: float | None = None
    note: str | None = None


class AccrualInputsPatch(BaseModel):
    period: str | None = None
    cells: list[AccrualInputPatch]


@router.patch("/accrual/inputs")
async def patch_accrual_inputs(
    payload: AccrualInputsPatch,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Lock (or clear) a deflator, a PLB rate, or a month's provisional flown.

    Refuses while the period is frozen: a frozen board is a booked number, and
    silently editing the inputs beneath it would make the snapshot unexplainable.
    """
    if payload.period:
        p = pa.resolve_period(payload.period)
        if await _frozen_for(db, current_user, p.key) is not None:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail=f"{p.label} is frozen. Re-open it before editing inputs.",
            )

    touched = 0
    for cell in payload.cells:
        keys = dict(
            tenant_id=current_user.tenant_id,
            created_by_id=current_user.id,
            airline_key=pa.norm_key(cell.airline_name),
            entity_key=pa.norm_key(cell.entity),
            channel_key=pa.norm_key(cell.channel),
            lob_key=pa.norm_key(cell.lob),
            ym=cell.ym,
        )
        row = (await db.execute(
            select(PlbAccrualInput).filter_by(**keys)
        )).scalar_one_or_none()

        sent = cell.model_fields_set
        if row is None:
            row = PlbAccrualInput(**keys)
            db.add(row)
        for field_name in ("deflator_pct", "plb_rate_pct", "manual_flown", "note"):
            if field_name in sent:
                setattr(row, field_name, getattr(cell, field_name))
        row.updated_by_id = current_user.id
        touched += 1

        # A row holding nothing is noise — and it would keep claiming the unique
        # slot, so the next "lock this" would update a tombstone instead of insert.
        if (row.deflator_pct is None and row.plb_rate_pct is None
                and row.manual_flown is None and not row.note):
            if row in db.new:
                db.expunge(row)
            else:
                await db.delete(row)

    await db.commit()
    return {"updated": touched}


class AirlineSettingPatch(BaseModel):
    airline_name: str
    entity: str | None = None
    channel: str | None = None
    flown_confirmed_through: date | None = None
    default_deflator_pct: float | None = None
    note: str | None = None


@router.patch("/accrual/settings")
async def patch_airline_settings(
    payload: list[AirlineSettingPatch],
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Set the 'Flown Confirmation Status' month (and an optional default deflator)
    for an airline × entity × channel."""
    for item in payload:
        keys = dict(
            tenant_id=current_user.tenant_id,
            created_by_id=current_user.id,
            airline_key=pa.norm_key(item.airline_name),
            entity_key=pa.norm_key(item.entity),
            channel_key=pa.norm_key(item.channel),
        )
        row = (await db.execute(
            select(PlbAirlineSetting).filter_by(**keys)
        )).scalar_one_or_none()
        if row is None:
            row = PlbAirlineSetting(**keys)
            db.add(row)
        sent = item.model_fields_set
        for field_name in ("flown_confirmed_through", "default_deflator_pct", "note"):
            if field_name in sent:
                setattr(row, field_name, getattr(item, field_name))
        # Stored as the first of the month so a month bucket compares cleanly.
        if row.flown_confirmed_through:
            row.flown_confirmed_through = row.flown_confirmed_through.replace(day=1)
        row.updated_by_id = current_user.id
    await db.commit()
    return {"updated": len(payload)}


# ── Freeze / re-open ─────────────────────────────────────────────────────────

class FreezeRequest(BaseModel):
    period: str | None = None
    basis: str = "auto"
    note: str | None = None


@router.post("/accrual/freeze", response_model=FrozenRead, status_code=status.HTTP_201_CREATED)
async def freeze_period(
    payload: FreezeRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Book this period's accrual: store the grid exactly as it stands.

    Deliberately refuses to overwrite an existing snapshot. Re-freezing would
    replace a booked figure in place with no trace of the old one; re-opening is
    an explicit, separate act.
    """
    p = pa.resolve_period(payload.period)
    if await _frozen_for(db, current_user, p.key) is not None:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"{p.label} is already frozen. Re-open it first.",
        )

    board = await pa.build_board(db, current_user.tenant_id, current_user.id, p, payload.basis)
    months = board["months"]
    grid = {
        "months": months,
        "basis": board["basis"],
        "data_quality": board["data_quality"],
        "rows": [
            {
                **{k: v for k, v in vars(r).items() if k != "months"},
                "plb_period_from": r.plb_period_from.isoformat() if r.plb_period_from else None,
                "plb_period_to": r.plb_period_to.isoformat() if r.plb_period_to else None,
                "flown_confirmed_through": (
                    r.flown_confirmed_through.isoformat() if r.flown_confirmed_through else None
                ),
                "months": {ym: vars(r.months[ym]) for ym in months if ym in r.months},
            }
            for r in board["rows"]
        ],
    }
    snap = PlbAccrualSnapshot(
        tenant_id=current_user.tenant_id,
        created_by_id=current_user.id,
        period_key=p.key,
        period_from=p.start,
        period_to=p.end,
        grid=grid,
        totals=board["totals"],
        row_count=len(board["rows"]),
        total_accrual=board["totals"]["accrual"],
        note=payload.note,
        frozen_by_id=current_user.id,
    )
    db.add(snap)
    await db.commit()
    await db.refresh(snap)
    return FrozenRead(
        period_key=snap.period_key, frozen_at=snap.frozen_at,
        total_accrual=_f(snap.total_accrual), row_count=snap.row_count, note=snap.note,
    )


@router.delete("/accrual/freeze/{period_key}", status_code=status.HTTP_204_NO_CONTENT)
async def reopen_period(
    period_key: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    snap = await _frozen_for(db, current_user, pa.resolve_period(period_key).key)
    if snap is None:
        raise HTTPException(status_code=404, detail="That period is not frozen.")
    await db.delete(snap)
    await db.commit()
    return Response(status_code=status.HTTP_204_NO_CONTENT)


@router.get("/accrual/snapshots", response_model=list[FrozenRead])
async def list_snapshots(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    rows = (await db.execute(
        select(PlbAccrualSnapshot).where(
            PlbAccrualSnapshot.tenant_id == current_user.tenant_id,
            PlbAccrualSnapshot.created_by_id == current_user.id,
        ).order_by(PlbAccrualSnapshot.period_from.desc())
    )).scalars().all()
    return [
        FrozenRead(
            period_key=s.period_key, frozen_at=s.frozen_at,
            total_accrual=_f(s.total_accrual), row_count=s.row_count, note=s.note,
        )
        for s in rows
    ]


# ── Excel export ─────────────────────────────────────────────────────────────

@router.get("/accrual/xlsx")
async def export_accrual_xlsx(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    period: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    basis: str = Query(default="auto", pattern="^(auto|travel|issue)$"),
    airline: list[str] | None = Query(default=None),
    entity: list[str] | None = Query(default=None),
    channel: list[str] | None = Query(default=None),
    lob: list[str] | None = Query(default=None),
    status_flag: list[str] | None = Query(default=None, alias="status"),
    search: str | None = Query(default=None),
):
    """The board as a workbook, laid out in the same column order as the source
    spreadsheet so it drops straight into an existing review process."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    p, board = await _board(db, current_user, period, date_from, date_to, basis)
    months = board["months"]
    rows = pa.filter_rows(board["rows"], airline, entity, channel, lob, status_flag, search)
    totals = pa.compute_totals(rows, months)

    wb = Workbook()
    ws = wb.active
    ws.title = p.key[:31]

    def _mon(ym: str) -> str:
        """'2026-04' -> 'Apr-26', the header the source workbook uses."""
        return f"{date(int(ym[:4]), int(ym[5:7]), 1):%b}-{ym[2:4]}"

    headers = [
        "Airline Name", "GDS/LCC", "Entity", "LOB", "PLB Period",
        "Flown Confirmation", "Basic", "Deflator Rate", "PLB rate",
        *[_mon(ym) for ym in months],
        f"{p.label} Final PLB", "Status",
    ]
    ws.append(headers)

    head_fill = PatternFill("solid", fgColor="1E3A5F")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = head_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Same two colours the manual sheet uses, now driven by computed status.
    red = PatternFill("solid", fgColor="FFC7CE")
    amber = PatternFill("solid", fgColor="FFEB9C")
    severity = {
        "EXPIRED_WITH_FLOWN": red, "NO_DEAL": red,
        "NO_RATE": amber, "NEEDS_SPLIT": amber, "EXPIRING": red, "UNCONFIRMED": amber,
    }

    for r in rows:
        ws.append([
            r.airline_name, r.channel, r.entity or "", r.lob or "", r.plb_period_label,
            r.flown_confirmed_through.strftime("%b-%y") if r.flown_confirmed_through else "",
            r.basis_label,
            round(r.deflator_pct / 100, 6),
            round(r.plb_rate_pct / 100, 6),
            *[r.months[ym].flown if ym in r.months else 0 for ym in months],
            r.accrual,
            ", ".join(r.status_flags),
        ])
        fill = severity.get(r.status)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    ws.append([
        "TOTAL", "", "", "", "", "", "", "", "",
        *[totals["by_month"].get(ym, 0) for ym in months],
        totals["accrual"], "",
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    money = "#,##0"
    for col in range(10, 10 + len(months) + 1):
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = money
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=8).number_format = "0.00%"
        ws.cell(row=row, column=9).number_format = "0.00%"

    widths = [26, 9, 10, 10, 18, 16, 14, 13, 10] + [15] * len(months) + [16, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "E2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="plb-accrual-{p.key}.xlsx"'},
    )


# ══════════════════════════════════════════════════════════════════════════════
# Overview
# ══════════════════════════════════════════════════════════════════════════════

class MonthPoint(BaseModel):
    ym: str
    label: str
    flown: float
    accrual: float


class AirlinePoint(BaseModel):
    airline: str
    flown: float
    accrual: float
    share_pct: float
    cumulative_pct: float


class EntityRow(BaseModel):
    entity: str
    accrual: float
    by_airline: dict[str, float]


class ExceptionGroup(BaseModel):
    code: str
    label: str
    count: int
    amount: float
    airlines: list[str]


class ActionCounts(BaseModel):
    pending_deal_approvals: int
    deals_awaiting_review: int
    statements_awaiting_commission: int
    unmatched_commission_rows: int


class OverviewResponse(BaseModel):
    period: PeriodRead
    totals: dict
    monthly: list[MonthPoint]
    by_airline: list[AirlinePoint]
    by_entity: list[EntityRow]
    entity_airlines: list[str]
    exceptions: list[ExceptionGroup]
    actions: ActionCounts
    data_quality: dict
    frozen: FrozenRead | None


_EXCEPTION_LABELS = {
    "EXPIRED_WITH_FLOWN": "Flown revenue with no active deal",
    "NO_DEAL": "Airlines with no PLB deal on file",
    "NO_RATE": "Zero rate on live volume",
    "NEEDS_SPLIT": "Flown not attributed to an entity",
    "EXPIRING": "Deals expiring within 90 days",
    "UNCONFIRMED": "Months the airline has not confirmed",
}


@router.get("/overview", response_model=OverviewResponse)
async def get_overview(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    period: str | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    basis: str = Query(default="auto", pattern="^(auto|travel|issue)$"),
    entity: list[str] | None = Query(default=None),
    channel: list[str] | None = Query(default=None),
    top: int = Query(default=10, ge=3, le=25),
):
    """KPI band, charts and exception counts — all derived from the same board the
    grid renders, so the two can never disagree."""
    tid, uid = current_user.tenant_id, current_user.id
    p, board = await _board(db, current_user, period, date_from, date_to, basis)
    months = board["months"]
    rows = pa.filter_rows(board["rows"], entities=entity, channels=channel)
    totals = pa.compute_totals(rows, months)

    monthly = [
        MonthPoint(
            ym=ym,
            label=f"{date(int(ym[:4]), int(ym[5:7]), 1):%b %y}",
            flown=totals["by_month"].get(ym, 0.0),
            accrual=totals["accrual_by_month"].get(ym, 0.0),
        )
        for ym in months
    ]

    # Pareto: the airlines carrying the accrual, and how few of them it takes.
    per_airline: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0])
    for r in rows:
        per_airline[r.airline_name][0] += r.accrual
        per_airline[r.airline_name][1] += r.flown_in_period
    ranked = sorted(per_airline.items(), key=lambda kv: -kv[1][0])[:top]
    grand = totals["accrual"] or 0.0
    by_airline: list[AirlinePoint] = []
    running = 0.0
    for name, (acc, flown) in ranked:
        running += acc
        by_airline.append(AirlinePoint(
            airline=name,
            flown=round(flown, 2),
            accrual=round(acc, 2),
            share_pct=round(acc / grand * 100, 2) if grand else 0.0,
            cumulative_pct=round(running / grand * 100, 2) if grand else 0.0,
        ))

    heat_airlines = [a.airline for a in by_airline[:6]]
    per_entity: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    entity_total: dict[str, float] = defaultdict(float)
    for r in rows:
        label = r.entity or "Unassigned"
        entity_total[label] += r.accrual
        if r.airline_name in heat_airlines:
            per_entity[label][r.airline_name] += r.accrual
    by_entity = [
        EntityRow(
            entity=e,
            accrual=round(entity_total[e], 2),
            by_airline={a: round(per_entity[e].get(a, 0.0), 2) for a in heat_airlines},
        )
        for e in sorted(entity_total, key=lambda x: -entity_total[x])
    ]

    exceptions: list[ExceptionGroup] = []
    for code, label in _EXCEPTION_LABELS.items():
        hits = [r for r in rows if code in r.status_flags]
        if not hits:
            continue
        # The number that matters differs by exception: leakage is measured in
        # flown revenue we are not claiming on, an expiring deal in the accrual
        # that is about to stop.
        amount = (
            sum(r.flown_total for r in hits)
            if code in ("EXPIRED_WITH_FLOWN", "NO_DEAL", "NO_RATE", "NEEDS_SPLIT")
            else sum(r.accrual for r in hits)
        )
        exceptions.append(ExceptionGroup(
            code=code, label=label, count=len(hits), amount=round(amount, 2),
            airlines=sorted({r.airline_name for r in hits})[:6],
        ))
    exceptions.sort(key=lambda e: pa.STATUS_ORDER.index(e.code))

    pending_approvals = (await db.execute(
        select(func.count(DealApproval.id)).where(
            DealApproval.status == ApprovalActionStatus.PENDING,
            DealApproval.submitted_by_id == uid,
        )
    )).scalar_one() or 0
    awaiting_review = (await db.execute(
        select(func.count(Deal.id)).where(
            Deal.tenant_id == tid, Deal.created_by_id == uid,
            Deal.direction == DealDirection.INBOUND,
            Deal.status == DealStatusType.PENDING_APPROVAL,
        )
    )).scalar_one() or 0
    awaiting_commission = (await db.execute(
        select(func.count(BspStatement.batch_id)).where(
            BspStatement.tenant_id == tid, BspStatement.created_by_id == uid,
            BspStatement.status == "completed",
            BspStatement.commission_status != "completed",
        )
    )).scalar_one() or 0
    unmatched_rows = (await db.execute(
        select(func.count(BspStatementRow.id)).where(
            BspStatementRow.tenant_id == tid, BspStatementRow.created_by_id == uid,
            BspStatementRow.commission_status == "unmatched",
        )
    )).scalar_one() or 0

    snap = await _frozen_for(db, current_user, p.key)
    return OverviewResponse(
        period=PeriodRead(key=p.key, label=p.label, **{"from": p.start}, to=p.end),
        totals=totals,
        monthly=monthly,
        by_airline=by_airline,
        by_entity=by_entity,
        entity_airlines=heat_airlines,
        exceptions=exceptions,
        actions=ActionCounts(
            pending_deal_approvals=pending_approvals,
            deals_awaiting_review=awaiting_review,
            statements_awaiting_commission=awaiting_commission,
            unmatched_commission_rows=unmatched_rows,
        ),
        data_quality=board["data_quality"],
        frozen=None if snap is None else FrozenRead(
            period_key=snap.period_key, frozen_at=snap.frozen_at,
            total_accrual=_f(snap.total_accrual), row_count=snap.row_count, note=snap.note,
        ),
    )


# ══════════════════════════════════════════════════════════════════════════════
# Income Summary — UNCHANGED. Owned by the /income-summary page, not this module.
# ══════════════════════════════════════════════════════════════════════════════

_DOM_VARIANTS = {"DOM", "DOMESTIC", "D"}
_INT_VARIANTS = {"INT", "INTL", "INTERNATIONAL", "I"}


class MonthlyBreakdown(BaseModel):
    month:       str
    commission:  float
    incentive:   float
    delta_comm:  float


class AirlineBreakdown(BaseModel):
    airline:    str
    commission: float
    incentive:  float
    delta_comm: float
    total:      float


class IncomeSummaryResponse(BaseModel):
    total:      float
    commission: float
    incentive:  float
    delta_comm: float
    monthly:    list[MonthlyBreakdown]
    by_airline: list[AirlineBreakdown]


class IncomeFiltersResponse(BaseModel):
    airlines:    list[str]
    segments:    list[str]
    class_types: list[str]


@router.get("/income-filters", response_model=IncomeFiltersResponse)
async def get_income_filters(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    airline: str | None = Query(default=None),
):
    a_res = await db.execute(
        select(Airline.name).where(Airline.is_active == True).order_by(Airline.name)  # noqa: E712
    )
    airlines = [r[0] for r in a_res.all()]

    segments = ["Domestic", "International"]

    ct_q = select(AirlineClassMaster.class_type).where(AirlineClassMaster.is_active == True).distinct()  # noqa: E712
    if airline:
        ct_q = ct_q.where(func.lower(AirlineClassMaster.airline_name) == airline.lower())
    ct_res = await db.execute(ct_q.order_by(AirlineClassMaster.class_type))
    class_types = [r[0] for r in ct_res.all()]

    return IncomeFiltersResponse(airlines=airlines, segments=segments, class_types=class_types)


@router.get("/income-summary", response_model=IncomeSummaryResponse)
async def get_income_summary(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    airline: str | None = Query(default=None),
    segment: str | None = Query(default=None),
    class_type: str | None = Query(default=None),
):
    tid = current_user.tenant_id
    filters = [
        UploadedTicket.tenant_id == tid,
        UploadedTicket.created_by_id == current_user.id,
    ]

    if airline:
        filters.append(UploadedTicket.airline_name == airline)

    if segment:
        seg_upper = segment.upper()
        if seg_upper in _DOM_VARIANTS:
            filters.append(func.upper(UploadedTicket.segment_type).in_(list(_DOM_VARIANTS)))
        elif seg_upper in _INT_VARIANTS:
            filters.append(func.upper(UploadedTicket.segment_type).in_(list(_INT_VARIANTS)))

    if class_type:
        cc_q = select(AirlineClassMaster.class_code).where(
            AirlineClassMaster.class_type == class_type,
            AirlineClassMaster.is_active == True,  # noqa: E712
        )
        if airline:
            cc_q = cc_q.where(func.lower(AirlineClassMaster.airline_name) == airline.lower())
        cc_res = await db.execute(cc_q)
        class_codes = [r[0] for r in cc_res.all()]
        if class_codes:
            filters.append(UploadedTicket.booking_class.in_(class_codes))

    tickets_res = await db.execute(select(UploadedTicket).where(*filters))
    tickets = tickets_res.scalars().all()

    total            = sum(_f(t.sell_fare) for t in tickets)
    total_commission = sum(_f(t.comm_sell) for t in tickets)
    total_incentive  = sum(_f(t.calculated_incentive) for t in tickets)
    total_delta_comm = sum(_f(t.comm_sell) - _f(t.calculated_incentive) for t in tickets)

    m_comm:  dict[str, float] = defaultdict(float)
    m_inc:   dict[str, float] = defaultdict(float)
    m_delta: dict[str, float] = defaultdict(float)
    for t in tickets:
        ym = _ym(t.ticket_date) or _ym(str(t.created_at)[:10] if t.created_at else None)
        if ym:
            m_comm[ym]  += _f(t.comm_sell)
            m_inc[ym]   += _f(t.calculated_incentive)
            m_delta[ym] += _f(t.comm_sell) - _f(t.calculated_incentive)
    all_months = sorted(set(m_comm) | set(m_inc) | set(m_delta))[-12:]
    monthly = [
        MonthlyBreakdown(
            month=ym,
            commission=round(m_comm[ym], 2),
            incentive=round(m_inc[ym], 2),
            delta_comm=round(m_delta[ym], 2),
        )
        for ym in all_months
    ]

    a_comm:  dict[str, float] = defaultdict(float)
    a_inc:   dict[str, float] = defaultdict(float)
    a_delta: dict[str, float] = defaultdict(float)
    for t in tickets:
        key = t.airline_name or t.airlines_code or "Unknown"
        a_comm[key]  += _f(t.comm_sell)
        a_inc[key]   += _f(t.calculated_incentive)
        a_delta[key] += _f(t.comm_sell) - _f(t.calculated_incentive)
    all_airlines = sorted(a_comm, key=lambda k: -(a_comm[k] + a_inc[k] + a_delta[k]))[:15]
    by_airline = [
        AirlineBreakdown(
            airline=k,
            commission=round(a_comm[k], 2),
            incentive=round(a_inc[k], 2),
            delta_comm=round(a_delta[k], 2),
            total=round(a_comm[k] + a_inc[k] + a_delta[k], 2),
        )
        for k in all_airlines
    ]

    return IncomeSummaryResponse(
        total=round(total, 2),
        commission=round(total_commission, 2),
        incentive=round(total_incentive, 2),
        delta_comm=round(total_delta_comm, 2),
        monthly=monthly,
        by_airline=by_airline,
    )
