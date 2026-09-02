"""LCC Detailed Statement — dedicated upload wizard + scalable storage.

Flow (see frontend LccUploadWizard):
    POST /lcc-detailed/extract   -> store file to GCS, detect headers, auto-map,
                                    create a `staged` batch, return preview + mapping
    POST /lcc-detailed/confirm   -> persist the user's mapping, flip to `pending`,
                                    enqueue the Celery ingest task (HTTP 202)
    GET  /lcc-detailed/status/{batch_id}  -> poll processing progress
    GET  /lcc-detailed/records   -> paginated typed rows + folded Taxes/Segments/SSR

Mounted at top-level `/lcc-detailed` (NOT under the generic `/statements/{slug}`
catch-all). Scoped per user: every query filters tenant_id + created_by_id.
"""
from __future__ import annotations

import asyncio
import io
import uuid
from datetime import date, datetime, timedelta

import pandas as pd
from fastapi import (
    APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile, status,
)
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import and_, column, delete, func, literal, or_, select, update
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import aliased

from app.config import settings
from app.database import get_db
from app.dependencies import get_current_user
from app.models.corporate import Corporate
from app.models.customer import Customer
from app.models.lcc_batch_airline_id import LccBatchAirlineId
from app.models.lcc_detailed import LccDetailed, LccDetailedBatch
from app.models.tenant_airline import TenantAirline
from app.models.user import User
from app.services import customer_resolver as cres
from app.services import gcs
from app.services import lcc_billing_projection as proj
from app.services import lcc_detailed_spec as spec
from app.services.lcc_airline_selection import resolve_for_upload
from app.services.lcc_statement import _clean, detect_format

router = APIRouter()

_MIN_MATCHED_COLUMNS = 3
_SAMPLE_ROWS = 50

# Drill-in filters arrive as `f.<field>` (and `f.<field>.from` / `.to` for a date range),
# the same wire protocol the generic statements router uses, so adding one is a spec edit
# rather than a signature change. Unknown fields are ignored, never rejected — a stale
# bookmark should degrade, not 400.
_FILTER_PREFIX = "f."
# Facet dropdowns past this are unusable anyway; the client falls back to typing.
_MAX_FACET_VALUES = 200


def _bucket() -> str:
    return settings.GCS_BSP_BUCKET_NAME or settings.GCS_TICKETS_BUCKET_NAME


def _scope(model, user: User):
    return (model.tenant_id == user.tenant_id, model.created_by_id == user.id)


# ── drill-in filtering ───────────────────────────────────────────────────────
_FILTER_BY_FIELD: dict[str, dict] = {f["field"]: f for f in spec.FILTERS}
# Resolved ONCE at import, so a typo in the spec is an AttributeError at startup rather
# than a 500 in production — and so no request path ever calls getattr on user input.
# `__segments__` is excluded: it has no column, see _segments_cond.
_FILTER_COLS = {
    f["field"]: getattr(LccDetailed, f["field"])
    for f in spec.FILTERS if f["field"] != spec.SEGMENTS_FILTER_FIELD
}
_SUMMARY_COLS = {f["field"]: getattr(LccDetailed, f["field"]) for f in spec.SUMMARY_FIELDS}
# The date columns are DateTime; departure_date is a real Date. Drives the end-of-day
# handling in _date_cond, which is the one place this distinction changes the answer.
_DATE_ONLY_FIELDS = {"departure_date"}


def _like(value: str) -> str:
    """`%value%` with LIKE wildcards escaped, so a literal `_` or `%` searches for itself.

    Payment numbers and promo codes routinely contain underscores, and `_` is LIKE's
    single-character wildcard — unescaped, the search silently over-matches. Pair with
    `.ilike(pattern, escape="\\\\")`.
    """
    esc = value.replace("\\", "\\\\").replace("%", r"\%").replace("_", r"\_")
    return f"%{esc}%"


def _as_date(raw: str):
    """`YYYY-MM-DD` — what <input type="date"> emits — or None if it won't parse."""
    try:
        return date.fromisoformat(raw.strip()[:10])
    except (ValueError, AttributeError):
        return None


def _segments_cond(value: str):
    """Contains-match over the rendered `route flight_no` of ANY leg in `segments`.

    An EXISTS over jsonb_array_elements rather than casting the whole JSONB to text: the
    cast matches the object KEYS too, so searching "route" or "leg" would return every
    row. The string matched here is exactly what _seg_str puts in the Segments cell, so
    what you type is what you see.
    """
    seg = func.jsonb_array_elements(LccDetailed.segments).table_valued(column("value", JSONB))
    leg = seg.c.value
    return and_(
        LccDetailed.segments.isnot(None),
        select(literal(1)).select_from(seg).where(
            func.concat_ws(" ", leg["route"].astext, leg["flight_no"].astext)
            .ilike(_like(value), escape="\\")
        ).correlate(LccDetailed).exists(),
    )


def _date_cond(field: str, bound: str, value: str):
    """One end of a date range, or None if the value doesn't parse."""
    d = _as_date(value)
    if d is None:
        return None
    col = _FILTER_COLS[field]
    if field in _DATE_ONLY_FIELDS:
        return col >= d if bound == "from" else col <= d
    # DateTime column: `<= to` compares against 00:00:00 and would drop everything
    # transacted later that same day, so the upper bound is the NEXT midnight,
    # exclusive. Bare comparisons, no ::date cast, so any index still applies.
    return (col >= datetime.combine(d, datetime.min.time()) if bound == "from"
            else col < datetime.combine(d + timedelta(days=1), datetime.min.time()))


def _filter_conds(request: Request | None) -> list:
    """Declared `f.<field>` params on the request → SQLAlchemy conditions.

    The field name is only ever a dict key into the spec's allowlist, and the column comes
    from the pre-resolved _FILTER_COLS. Anything unknown, blank or unparseable is skipped
    rather than rejected — a stale bookmark should degrade, not 400.
    """
    if request is None:
        return []
    conds: list = []
    for key, raw in request.query_params.multi_items():
        if not key.startswith(_FILTER_PREFIX):
            continue
        field, _, bound = key[len(_FILTER_PREFIX):].partition(".")
        f = _FILTER_BY_FIELD.get(field)
        value = (raw or "").strip()
        if not f or not value:
            continue

        if f["type"] == "daterange":
            if bound in ("from", "to"):
                cond = _date_cond(field, bound, value)
                if cond is not None:
                    conds.append(cond)
        elif bound:
            continue          # `f.name1.from` on a non-daterange filter is malformed
        elif field == spec.SEGMENTS_FILTER_FIELD:
            conds.append(_segments_cond(value))
        elif f["type"] == "select":
            if f.get("options"):
                # Statically declared Yes/No over the `international` boolean.
                b = {"yes": True, "no": False}.get(value.lower())
                if b is not None:
                    conds.append(_FILTER_COLS[field].is_(b))
            else:
                conds.append(_FILTER_COLS[field] == value)
        else:
            conds.append(_FILTER_COLS[field].ilike(_like(value), escape="\\"))
    return conds


def _record_conds(user: User, batch_id: str | None, request: Request | None) -> list:
    """Scope + batch + whichever declared `f.<field>` filters the request carries."""
    conds = [*_scope(LccDetailed, user)]
    if batch_id:
        conds.append(LccDetailed.batch_id == batch_id)
    return conds + _filter_conds(request)


async def _tenant_airlines(
    db: AsyncSession, user: User, tenant_airline_ids: list[int]
) -> list[TenantAirline]:
    """Resolve the user's Airline Master entries for an upload.

    An LCC export carries no carrier, so this selection is the ONLY thing that
    identifies the airline. A statement may cover SEVERAL of the user's ids for that
    carrier — but not ids from two carriers; see services/lcc_airline_selection.py.
    """
    try:
        return await resolve_for_upload(db, user.tenant_id, tenant_airline_ids)
    except ValueError as exc:
        # MixedAirlineSelection and UnknownAirlineId are both ValueErrors and both
        # carry a message written for the user.
        raise HTTPException(status_code=400, detail=str(exc))


async def _stamp_airline(
    db: AsyncSession, batch: LccDetailedBatch, tas: list[TenantAirline]
) -> None:
    """Put the selection on the batch: the scalar columns from the primary (first)
    id, and the full set in the link table.

    The scalar airline_* columns stay single-valued and correct because every id in
    the selection shares one carrier — that is the whole reason for that rule.
    """
    primary = tas[0]
    batch.tenant_airline_id = primary.id
    batch.airline_id = primary.airline_id
    batch.airline_name = primary.airline_name
    batch.airline_code = primary.airline_code
    batch.airline_ref_id = primary.ref_id

    # Replace rather than merge: re-declaring the airline on a batch means "these are
    # the ids now", so an id dropped from the selection must not linger.
    await db.execute(
        delete(LccBatchAirlineId).where(LccBatchAirlineId.batch_id == batch.batch_id)
    )
    for ta in tas:
        db.add(LccBatchAirlineId(batch_id=batch.batch_id, tenant_airline_id=ta.id))


async def _batch_airline_ids(
    db: AsyncSession, batch_ids: list[str]
) -> dict[str, list[tuple[int, str]]]:
    """`{batch_id: [(tenant_airline_id, ref_id), ...]}` for a whole page of batches in
    ONE query — the batches list renders every row's id set, so a query per row would
    be an N+1 on the busiest screen in this view."""
    if not batch_ids:
        return {}
    rows = (await db.execute(
        select(LccBatchAirlineId.batch_id, TenantAirline.id, TenantAirline.ref_id)
        .join(TenantAirline, TenantAirline.id == LccBatchAirlineId.tenant_airline_id)
        .where(LccBatchAirlineId.batch_id.in_(batch_ids))
        .order_by(TenantAirline.ref_id)
    )).all()
    out: dict[str, list[tuple[int, str]]] = {}
    for batch_id, ta_id, ref_id in rows:
        out.setdefault(batch_id, []).append((ta_id, ref_id))
    return out


def _progress_pct(b: LccDetailedBatch) -> int:
    if b.status == "completed":
        return 100
    if b.total_rows:
        return min(100, int(b.processed_rows * 100 / b.total_rows))
    return 0


def _read_df(content: bytes, filename: str, header_row: int, nrows: int | None = None) -> pd.DataFrame:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        # index_col=False: these exports end each data row with a trailing delimiter, so
        # pandas would otherwise treat the first column as a row index and shift every
        # value left (garbage in every typed column). Force positional column alignment.
        return pd.read_csv(io.BytesIO(content), dtype=str, sep=None, engine="python",
                           header=header_row, index_col=False, nrows=nrows)
    if name.endswith(".xls"):
        return pd.read_excel(io.BytesIO(content), dtype=str, header=header_row, nrows=nrows)
    return pd.read_excel(io.BytesIO(content), dtype=str, engine="openpyxl", header=header_row, nrows=nrows)


def _detect_df(content: bytes, filename: str) -> tuple[pd.DataFrame, int, int]:
    """Pick the header row (0/1/2) whose columns recognise the most standard columns.
    Returns (df, header_row, matched_count)."""
    best: tuple[int, pd.DataFrame, int] | None = None
    last_err: Exception | None = None
    for hr in (0, 1, 2):
        try:
            df = _read_df(content, filename, hr)
        except Exception as e:  # noqa: BLE001 — unreadable at this offset, try the next
            last_err = e
            continue
        df.dropna(how="all", inplace=True)
        cols = [str(c) for c in df.columns]
        _, matched, _ = spec.suggest_mapping(cols)
        if best is None or matched > best[0]:
            best = (matched, df, hr)
    if best is None:
        raise HTTPException(status_code=400, detail=f"Could not read the file: {last_err}. Upload a valid .xlsx, .xls or .csv.")
    matched, df, hr = best
    if matched < _MIN_MATCHED_COLUMNS:
        sample = ", ".join(spec.LCC_STANDARD_HEADERS[:5])
        raise HTTPException(
            status_code=400,
            detail=f"This doesn't look like an LCC Detailed statement — only {matched} known "
                   f"column(s) were recognised. Expected headers such as: {sample}…",
        )
    return df, hr, matched


def _seg_str(s: dict) -> str:
    return f"{s.get('route') or ''} {s.get('flight_no') or ''}".strip()


def _disp(v) -> str:
    if v is None:
        return ""
    if isinstance(v, bool):
        return "Yes" if v else "No"
    return str(v)


def _money_str(value) -> str | None:
    """Serialize a summed Decimal as a string — a float round-trip would lose paise."""
    return None if value is None else format(value.normalize(), "f")


# ── mapping / metadata ───────────────────────────────────────────────────────
@router.get("/standard-columns")
async def standard_columns(current_user: User = Depends(get_current_user)):
    """Grouped standard column spec for the mapping UI."""
    return {"groups": spec.grouped_columns(), "total": len(spec.LCC_STANDARD_COLUMNS)}


@router.get("/template")
async def download_template(current_user: User = Depends(get_current_user)):
    """Blank .xlsx with the exact 129 standard headers."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "LCC Detailed Template"
    ws.append(spec.LCC_STANDARD_HEADERS)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="lcc_detailed_template.xlsx"'},
    )


# ── extract → confirm ────────────────────────────────────────────────────────
@router.post("/extract")
async def extract(
    file: UploadFile = File(...),
    tenant_airline_ids: list[int] = Form(default=[]),
    tenant_airline_id: int | None = Form(default=None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Store the file, detect headers, auto-map, and stage a batch. Nothing is
    ingested yet — the user reviews the mapping, then calls /confirm.

    The airline selection is required and resolved here rather than at /confirm, so a
    batch is never in a state where rows could be ingested without an airline. Several
    ids may be sent — one statement usually covers more than one of the user's logins
    for that carrier — but they must all be the same airline.

    `tenant_airline_id` (singular) is still accepted so a client that has not picked up
    the new field cannot suddenly 422 mid-upload."""
    picked = list(tenant_airline_ids)
    if tenant_airline_id is not None and tenant_airline_id not in picked:
        picked.append(tenant_airline_id)
    tas = await _tenant_airlines(db, current_user, picked)
    ta = tas[0]
    content = await file.read()
    max_bytes = settings.MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(status_code=413, detail=f"File exceeds {settings.MAX_UPLOAD_SIZE_MB} MB limit.")
    if not content:
        raise HTTPException(status_code=400, detail="The uploaded file is empty.")

    df, header_row, matched = _detect_df(content, file.filename or "")
    cols = [str(c) for c in df.columns]
    total_rows = int(len(df))
    suggested_mapping, matched_columns, is_template_match = spec.suggest_mapping(cols)
    source_format = detect_format(cols)

    sample_df = df.head(_SAMPLE_ROWS)
    sample_rows = [
        {str(c): _clean(row[c]) for c in cols}
        for _, row in sample_df.iterrows()
    ]

    batch_id = str(uuid.uuid4())
    # Store the source file — REQUIRED (the worker re-downloads it). Hard-fail otherwise.
    blob_name = f"lcc-detailed/{current_user.tenant_id}/{batch_id}/{file.filename}"
    try:
        await gcs.upload_bytes(content, blob_name, file.content_type or "application/octet-stream", _bucket())
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=502, detail=f"Could not store the file: {exc}")

    batch = LccDetailedBatch(
        batch_id=batch_id,
        tenant_id=current_user.tenant_id,
        created_by_id=current_user.id,
        source_file=file.filename,
        file_url=blob_name,
        source_format=source_format,
        header_row=header_row,
        column_map=suggested_mapping,
        status="staged",
        total_rows=total_rows,
        processed_rows=0,
        matched_columns=matched_columns,
        uploaded_at=datetime.utcnow(),
    )
    db.add(batch)
    # After db.add: _stamp_airline writes the link rows, whose FK is the batch.
    await _stamp_airline(db, batch, tas)
    await db.commit()

    return {
        "batch_id": batch_id,
        "file_name": file.filename,
        "airline_name": ta.airline_name,
        "airline_code": ta.airline_code,
        "airline_ref_id": ta.ref_id,
        "airline_ref_ids": [t.ref_id for t in tas],
        "tenant_airline_ids": [t.id for t in tas],
        "total_rows": total_rows,
        "header_row": header_row,
        "source_format": source_format,
        "xls_columns": cols,
        "suggested_mapping": suggested_mapping,
        "sample_rows": sample_rows,
        "is_template_match": is_template_match,
        "matched_columns": matched_columns,
        "standard_total": len(spec.LCC_STANDARD_COLUMNS),
    }


class ConfirmPayload(BaseModel):
    batch_id: str
    column_map: dict[str, str]
    expected_rows: int | None = None   # user-declared expected record count (optional)


@router.post("/confirm", status_code=status.HTTP_202_ACCEPTED)
async def confirm(
    payload: ConfirmPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Persist the confirmed mapping and enqueue background ingestion."""
    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == payload.batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Upload not found.")
    if batch.status != "staged":
        raise HTTPException(status_code=409, detail=f"This upload is already '{batch.status}'.")

    # Keep only mappings that point at a real standard field.
    valid_fields = {c["field"] for c in spec.LCC_STANDARD_COLUMNS}
    clean_map = {k: v for k, v in payload.column_map.items() if k in valid_fields and v}
    batch.column_map = clean_map
    batch.matched_columns = len(clean_map)
    batch.expected_rows = payload.expected_rows if (payload.expected_rows or 0) > 0 else None
    batch.status = "pending"
    batch.error = None
    await db.commit()   # commit BEFORE enqueue so the worker can't 404 on its own batch

    from app.workers.lcc_tasks import ingest_lcc_detailed
    ingest_lcc_detailed.delay(payload.batch_id, current_user.tenant_id, current_user.id)

    return {"batch_id": payload.batch_id, "status": "pending", "total_rows": batch.total_rows}


@router.get("/status/{batch_id}")
async def get_status(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Upload not found.")
    return {
        "batch_id": batch.batch_id,
        "status": batch.status,
        "total_rows": batch.total_rows,
        "processed_rows": batch.processed_rows,
        "expected_rows": batch.expected_rows,
        "progress_pct": _progress_pct(batch),
        "matched_columns": batch.matched_columns,
        "error": batch.error,
        "source_file": batch.source_file,
        "completed_at": batch.completed_at,
    }


# ── batches / records ────────────────────────────────────────────────────────
@router.get("/batches")
async def list_batches(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """One row per upload, straight from the header table (no GROUP BY)."""
    rows = (await db.execute(
        select(LccDetailedBatch)
        .where(*_scope(LccDetailedBatch, current_user))
        .order_by(LccDetailedBatch.uploaded_at.desc())
    )).scalars().all()
    links = await _batch_airline_ids(db, [b.batch_id for b in rows])
    return [
        {
            "batch_id": b.batch_id,
            "source_file": b.source_file,
            "uploaded_at": b.uploaded_at,
            "completed_at": b.completed_at,
            "status": b.status,
            "total_rows": b.total_rows,
            "processed_rows": b.processed_rows,
            "expected_rows": b.expected_rows,
            "progress_pct": _progress_pct(b),
            "row_count": b.processed_rows if b.status == "completed" else b.total_rows,
            "has_file": bool(b.file_url),
            "created_by_name": current_user.full_name,
            "airline_name": b.airline_name,
            "airline_code": b.airline_code,
            "airline_ref_id": b.airline_ref_id,
            "tenant_airline_id": b.tenant_airline_id,
            # The full set this upload covers. Falls back to the primary columns for a
            # batch that predates the link table and was never re-saved.
            "airline_ref_ids": (
                [ref for _, ref in links[b.batch_id]] if b.batch_id in links
                else ([b.airline_ref_id] if b.airline_ref_id else [])
            ),
            "tenant_airline_ids": (
                [tid for tid, _ in links[b.batch_id]] if b.batch_id in links
                else ([b.tenant_airline_id] if b.tenant_airline_id else [])
            ),
            "billable_rows": b.billable_rows,
            "resolved_rows": b.resolved_rows,
            "unresolved_rows": b.unresolved_rows,
            "projected_rows": b.projected_rows,
            "resolution_status": b.resolution_status,
        }
        for b in rows
    ]


class BatchAirlinePayload(BaseModel):
    # Several ids may be sent — one statement usually covers more than one of the
    # user's logins for that carrier. The singular field is still accepted so an
    # older client keeps working; see extract() for the same pairing.
    tenant_airline_ids: list[int] = []
    tenant_airline_id: int | None = None

    def picked(self) -> list[int]:
        out = list(self.tenant_airline_ids)
        if self.tenant_airline_id is not None and self.tenant_airline_id not in out:
            out.append(self.tenant_airline_id)
        return out


@router.patch("/batches/{batch_id}/airline")
async def set_batch_airline(
    batch_id: str,
    payload: BatchAirlinePayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Set (or correct) the airline on an already-uploaded batch.

    Batches imported before the airline was captured have none, and the rows parsed
    correctly — only the carrier was missing. So this stamps the header and issues one
    indexed bulk UPDATE over the rows rather than re-uploading or re-ingesting."""
    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Upload not found.")

    tas = await _tenant_airlines(db, current_user, payload.picked())
    ta = tas[0]
    await _stamp_airline(db, batch, tas)

    result = await db.execute(
        update(LccDetailed)
        .where(LccDetailed.batch_id == batch_id, *_scope(LccDetailed, current_user))
        .values(
            airline_id=ta.airline_id,
            airline_name=ta.airline_name,
            airline_code=ta.airline_code,
        )
    )
    await db.commit()

    return {
        "batch_id": batch_id,
        "airline_name": ta.airline_name,
        "airline_code": ta.airline_code,
        "airline_ref_id": ta.ref_id,
        "airline_ref_ids": [t.ref_id for t in tas],
        "tenant_airline_ids": [t.id for t in tas],
        "rows_updated": result.rowcount or 0,
    }


# ── billing: resolve each row to a customer / corporate ──────────────────────
# An LCC export names no customer, only a passenger per row. These endpoints resolve
# that passenger against the Customer master so the rows can later be projected into
# `uploaded_tickets`, the one table Customer/Corporate Billing reads.

MAX_GAP_GROUPS = 50          # mirrors api/v1/bsp_commission.py:54
_BILL_PARTY_TYPES = ("corporate", "direct")

# How many hand-picked rows one call may carry. 500 matches bsp_commission's
# MAX_INLINE_ROWS; the cap exists because project_batch flushes per inserted row, so a
# selection is a quick action rather than a full run in disguise. The two are kept EQUAL
# so a "select all matching" can always be posted back in one call.
MAX_SEND_ROWS = 500
MAX_BILLING_SELECT_IDS = 500


def _bill_kind(total) -> str:
    """Classify a row from `total`, NOT `base_fare`.

    A fee-only row (no-show, seat, PSFR/UDFR re-charge) has base_fare 0 but real money
    in total, and a change fee can exceed a fare refund and flip the row's sign — on
    one real 203-row statement, classifying on base_fare drops 24 rows worth ₹13,581
    and mis-signs another. `total = 0` is a payment movement: money moved between
    accounts, no fare, nothing to bill.
    """
    if total is None or total == 0:
        return "payment"
    return "sale" if total > 0 else "refund"


async def _resolve_bill_party(
    db: AsyncSession, user: User,
    customer_type: str | None, customer_id: int | None, corporate_id: int | None,
) -> tuple[str | None, int | None, int | None]:
    """Authorise a party against the master rather than trusting the ids sent.

    Mirrors tickets.py::_resolve_customer_party. Returns the normalised triple, with
    the ids that do not belong to the chosen type nulled — the same discipline as
    frontend lib/customerType.ts::buildTagPayload, enforced server-side so a stale id
    from a previously chosen type can never travel attached to the wrong one.
    """
    ct = (customer_type or "").strip().lower() or None
    if ct is None:
        return None, None, None
    if ct not in _BILL_PARTY_TYPES:
        raise HTTPException(
            status_code=400,
            detail=("An LCC statement bills a customer or a corporate. Agency billing "
                    "claims tickets through their statement, not through this link."),
        )

    if ct == "corporate":
        if not corporate_id:
            raise HTTPException(status_code=400, detail="Pick a corporate.")
        row = (await db.execute(select(Corporate).where(
            Corporate.id == corporate_id,
            Corporate.tenant_id == user.tenant_id,
            Corporate.created_by_id == user.id,
        ))).scalar_one_or_none()
        if not row:
            raise HTTPException(status_code=400, detail=f"Corporate id {corporate_id} is not in your corporates.")
        # A customer may also be named, when the row resolved to an employee.
        cust_id = None
        if customer_id:
            cust = (await db.execute(select(Customer).where(
                Customer.id == customer_id,
                Customer.tenant_id == user.tenant_id,
                Customer.created_by_id == user.id,
            ))).scalar_one_or_none()
            if not cust:
                raise HTTPException(status_code=400, detail=f"Customer id {customer_id} is not in your customers.")
            cust_id = cust.id
        return "corporate", cust_id, row.id

    if not customer_id:
        raise HTTPException(status_code=400, detail="Pick a customer.")
    cust = (await db.execute(select(Customer).where(
        Customer.id == customer_id,
        Customer.tenant_id == user.tenant_id,
        Customer.created_by_id == user.id,
    ))).scalar_one_or_none()
    if not cust:
        raise HTTPException(status_code=400, detail=f"Customer id {customer_id} is not in your customers.")
    # A customer who belongs to a corporate is an employee — carry the corporate too,
    # because corporates.py defines a corporate's tickets as its employees' tickets.
    return ("corporate" if cust.corporate_id else "direct"), cust.id, cust.corporate_id


async def _owned_batch(batch_id: str, db: AsyncSession, current_user: User) -> LccDetailedBatch:
    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Upload not found.")
    return batch


async def _status_counts(db: AsyncSession, batch_id: str) -> dict[str, int]:
    """`{bill_status: n}` for one batch — the chips' source, and _recount's."""
    rows = (await db.execute(
        select(LccDetailed.bill_status, func.count())
        .where(LccDetailed.batch_id == batch_id)
        .group_by(LccDetailed.bill_status)
    )).all()
    return {s: n for s, n in rows}


async def _billing_state_counts(db: AsyncSession, batch_id: str) -> dict[str, int]:
    """`{billing_state: n}` for one batch, over the same LEFT JOIN billing-rows uses.

    Every state in one round trip via COUNT(*) FILTER (WHERE …) rather than seven
    queries. Scoped exactly like _status_counts — on batch_id alone, because the caller
    has already proved the batch is theirs through _owned_batch.

    Deliberately NOT narrowed by the caller's filters: this drives header counts, which
    must hold still while you page and search rather than re-describing the filter you
    just typed.
    """
    from app.models.uploaded_ticket import UploadedTicket

    T = aliased(UploadedTicket)
    row = (await db.execute(
        select(*[
            func.count().filter(proj.billing_state_cond(s, T)).label(s)
            for s in proj.BILLING_STATES
        ])
        .select_from(LccDetailed)
        .outerjoin(T, T.id == LccDetailed.projected_ticket_id)
        .where(LccDetailed.batch_id == batch_id)
    )).one()
    return {s: getattr(row, s) or 0 for s in proj.BILLING_STATES}


async def _recount(db: AsyncSession, batch: LccDetailedBatch) -> None:
    """Refresh the batch's billing counters from its rows."""
    billable = (cres.RESOLVED, cres.DEFAULTED, cres.OVERRIDDEN)
    counts = await _status_counts(db, batch.batch_id)
    batch.resolved_rows = sum(counts.get(s, 0) for s in billable)
    batch.unresolved_rows = sum(
        n for s, n in counts.items() if s not in billable and s != cres.EXCLUDED
    )
    batch.billable_rows = batch.resolved_rows + batch.unresolved_rows
    batch.projected_rows = await db.scalar(
        select(func.count()).select_from(LccDetailed)
        .where(LccDetailed.batch_id == batch.batch_id,
               LccDetailed.projected_ticket_id.isnot(None))
    ) or 0


class ResolvePayload(BaseModel):
    # A human's pick is the most authoritative thing on the row, so a re-run keeps it
    # unless the caller explicitly says otherwise.
    reset_overrides: bool = False


@router.post("/batches/{batch_id}/resolve-customers")
async def resolve_customers(
    batch_id: str,
    payload: ResolvePayload | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Classify every row and match its passenger against the Customer master."""
    payload = payload or ResolvePayload()
    batch = await _owned_batch(batch_id, db, current_user)
    if batch.status != "completed":
        raise HTTPException(status_code=409, detail="This upload is still importing.")

    index = await cres.CustomerIndex.load(
        db, tenant_id=current_user.tenant_id, created_by_id=current_user.id
    )
    rows = (await db.execute(
        select(LccDetailed.id, LccDetailed.name1, LccDetailed.total, LccDetailed.bill_status)
        .where(LccDetailed.batch_id == batch_id, *_scope(LccDetailed, current_user))
    )).all()

    now = datetime.utcnow()
    default_set = bool(batch.default_customer_type)
    updates: list[dict] = []
    summary: dict[str, int] = {}

    for row_id, name1, total, current_status in rows:
        kind = _bill_kind(total)

        if not payload.reset_overrides and current_status == cres.OVERRIDDEN and kind != "payment":
            summary[cres.OVERRIDDEN] = summary.get(cres.OVERRIDDEN, 0) + 1
            updates.append({"id": row_id, "bill_kind": kind})
            continue

        if kind == "payment":
            m = cres.CustomerMatch(status=cres.EXCLUDED, note=cres.REASON[cres.EXCLUDED])
        else:
            m = index.resolve(name1)
            if m.status == cres.UNRESOLVED and default_set:
                # The batch fallback is the primary path in practice: an LCC export's
                # passengers rarely overlap the Customer master at all.
                m = cres.CustomerMatch(
                    status=cres.DEFAULTED,
                    customer_id=batch.default_customer_id,
                    corporate_id=batch.default_corporate_id,
                    customer_type=batch.default_customer_type,
                    display_name=(name1 or "").strip(),
                    note="Billed to this upload's default party.",
                )

        summary[m.status] = summary.get(m.status, 0) + 1
        updates.append({
            "id": row_id,
            "bill_kind": kind,
            "bill_status": m.status,
            "bill_customer_type": m.customer_type,
            "bill_customer_id": m.customer_id,
            "bill_corporate_id": m.corporate_id,
            "bill_match_reason": m.note,
            "resolved_at": now,
            "resolved_by_id": current_user.id,
        })

    if updates:
        await db.execute(update(LccDetailed), updates)

    batch.resolution_status = "projected" if batch.resolution_status == "projected" else "resolved"
    await _recount(db, batch)
    await db.commit()

    return await _billing_summary(db, batch, customers_in_scope=len(index), summary=summary)


async def _billing_summary(
    db: AsyncSession, batch: LccDetailedBatch, *,
    customers_in_scope: int, summary: dict[str, int] | None = None,
) -> dict:
    """The worklist's header, in one shape.

    Returned by both `resolve-customers` (which has just recomputed `summary` in memory)
    and the read-only `billing-summary`, so the frontend has one type and one setter for
    the chips no matter which call refreshed them.
    """
    return {
        "batch_id": batch.batch_id,
        "customers_in_scope": customers_in_scope,
        "summary": summary if summary is not None else await _status_counts(db, batch.batch_id),
        "state_counts": await _billing_state_counts(db, batch.batch_id),
        "billable_rows": batch.billable_rows,
        "resolved_rows": batch.resolved_rows,
        "unresolved_rows": batch.unresolved_rows,
        "projected_rows": batch.projected_rows,
        "resolution_status": batch.resolution_status,
    }


@router.get("/batches/{batch_id}/billing-summary")
async def billing_summary(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """What `resolve-customers` returns, WITHOUT resolving anything.

    The worklist needs these counts every time it opens and after every edit. Getting
    them from the POST meant that merely looking at the screen re-matched every row —
    which, on an upload already sent to billing, could silently re-point a row billing
    already sees. This is the read-only way to ask.

    `customers_in_scope` is a plain count rather than len(CustomerIndex): its one
    consumer only tests it against zero, to say "your Customer master is empty".
    """
    batch = await _owned_batch(batch_id, db, current_user)
    in_scope = await db.scalar(
        select(func.count()).select_from(Customer).where(*_scope(Customer, current_user))
    ) or 0
    return await _billing_summary(db, batch, customers_in_scope=in_scope)


_BILLING_STATE_PATTERN = "^(" + "|".join((*proj.BILLING_STATES, "sendable")) + ")$"


@router.get("/batches/{batch_id}/billing-rows")
async def list_billing_rows(
    batch_id: str,
    status_filter: str | None = Query(None, alias="status"),
    billing_state: str | None = Query(None, pattern=_BILLING_STATE_PATTERN),
    q: str | None = Query(None, max_length=100),
    kind: str | None = Query(None),
    ids_only: bool = Query(False),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """The resolution worklist, paginated like /records.

    Two independent filters, because they answer two different questions: `status` is
    the row's MATCH status (does it have a party?) and `billing_state` is where it has
    got to on its way into billing. A row can be "Set by you" and "Ready to send" at the
    same time, so these are never merged into one param.

    `ids_only=true` returns just the matching ids, capped, for the "select all N
    matching" affordance. It lives on this endpoint rather than a sibling so the ids can
    only ever come from the same predicate builder as the visible page.
    """
    from app.models.uploaded_ticket import UploadedTicket

    await _owned_batch(batch_id, db, current_user)

    # The join is 1:1 — T.id is a primary key and projected_ticket_id is a FK to it — so
    # the COUNT over base.subquery() below cannot fan out. Do not "fix" it with DISTINCT.
    # UploadedTicket needs no _scope of its own: it is reachable only through
    # projected_ticket_id, which this same scoped pipeline is the only writer of.
    T = aliased(UploadedTicket)
    base = (select(LccDetailed, T)
            .outerjoin(T, T.id == LccDetailed.projected_ticket_id)
            .where(LccDetailed.batch_id == batch_id, *_scope(LccDetailed, current_user)))
    if status_filter:
        base = base.where(LccDetailed.bill_status == status_filter)
    if kind:
        base = base.where(LccDetailed.bill_kind == kind)
    if billing_state:
        base = base.where(proj.billing_state_cond(billing_state, T))
    if q and q.strip():
        # One box over both identities a user has to hand. _like escapes the LIKE
        # wildcards, which matters here: PNRs and payment refs carry underscores.
        pattern = _like(q.strip())
        base = base.where(or_(
            LccDetailed.name1.ilike(pattern, escape="\\"),
            LccDetailed.record_locator.ilike(pattern, escape="\\"),
        ))

    # Narrowed to the id before counting: the subquery would otherwise carry all ~130
    # columns of both tables for no reason. The join is 1:1 on a PK so the count is
    # unaffected either way.
    total = await db.scalar(
        select(func.count()).select_from(base.with_only_columns(LccDetailed.id).subquery())
    ) or 0

    if ids_only:
        ids = (await db.execute(
            base.with_only_columns(LccDetailed.id)
            .order_by(LccDetailed.id.asc()).limit(MAX_BILLING_SELECT_IDS)
        )).scalars().all()
        return {"total": total, "ids": list(ids), "truncated": total > len(ids)}

    pairs = (await db.execute(
        base.order_by(LccDetailed.id.asc()).limit(limit).offset(offset)
    )).all()
    rows = [r for r, _t in pairs]
    tickets = {r.id: t for r, t in pairs}

    # Resolve the party names for display in one query rather than per row. The id sets
    # carry the TICKET's party as well as the row's, so a stale row can name both.
    cust_ids = {r.bill_customer_id for r in rows if r.bill_customer_id}
    corp_ids = {r.bill_corporate_id for r in rows if r.bill_corporate_id}
    cust_ids |= {t.customer_id for t in tickets.values() if t is not None and t.customer_id}
    corp_ids |= {t.corporate_id for t in tickets.values() if t is not None and t.corporate_id}
    cust_names = dict((await db.execute(
        select(Customer.id, func.concat(Customer.first_name, " ", func.coalesce(Customer.last_name, "")))
        .where(Customer.id.in_(cust_ids or {-1}))
    )).all())
    corp_names = dict((await db.execute(
        select(Corporate.id, Corporate.company).where(Corporate.id.in_(corp_ids or {-1}))
    )).all())

    def _party_name(ct: str | None, cust_id: int | None, corp_id: int | None):
        if ct == "corporate":
            return corp_names.get(corp_id)
        return (cust_names.get(cust_id) or "").strip() or None

    def _row(r):
        t = tickets.get(r.id)
        state = proj.billing_state(r, t)
        return {
            "id": r.id,
            "passenger": r.name1,
            "record_locator": r.record_locator,
            "transaction_date": r.transaction_date,
            "departure_date": r.departure_date,
            "total": float(r.total) if r.total is not None else None,
            "base_fare": float(r.base_fare) if r.base_fare is not None else None,
            "bill_kind": r.bill_kind,
            "bill_status": r.bill_status,
            "bill_customer_type": r.bill_customer_type,
            "bill_customer_id": r.bill_customer_id,
            "bill_corporate_id": r.bill_corporate_id,
            "party_name": _party_name(r.bill_customer_type, r.bill_customer_id, r.bill_corporate_id),
            "customer_name": (cust_names.get(r.bill_customer_id) or "").strip() or None,
            "bill_match_reason": r.bill_match_reason,
            "projected_ticket_id": r.projected_ticket_id,
            "billing_state": state,
            "billing_id": t.billing_id if t is not None else None,
            # Only for a stale row, and it is the whole point of that state: the screen
            # can say "sent as X, now billed to Y" instead of an unexplained warning.
            "billed_party_name": (
                _party_name(t.customer_type, t.customer_id, t.corporate_id)
                if state == "stale" else None
            ),
            "sendable": state in proj.SENDABLE_STATES,
        }

    # No state_counts here on purpose: they cover the whole batch, so recomputing them
    # per page would put a full-batch aggregate behind every search keystroke. The
    # header gets them from billing-summary, which is called after mutations instead.
    return {
        "total": total, "limit": limit, "offset": offset,
        "rows": [_row(r) for r in rows],
    }


@router.get("/batches/{batch_id}/billing-gaps")
async def billing_gaps(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Rows grouped by (status, reason), with a few example passengers each.

    Structurally the same as bsp_commission's /gaps. This only collapses because
    `bill_match_reason` is identical per gap type — see customer_resolver.REASON.
    """
    await _owned_batch(batch_id, db, current_user)
    billable = (cres.RESOLVED, cres.DEFAULTED, cres.OVERRIDDEN)

    rows = (await db.execute(
        select(
            LccDetailed.bill_status,
            LccDetailed.bill_match_reason,
            func.count().label("n"),
            func.array_agg(func.coalesce(LccDetailed.name1, "—")).label("samples"),
        )
        .where(LccDetailed.batch_id == batch_id, *_scope(LccDetailed, current_user),
               LccDetailed.bill_status.notin_(billable))
        .group_by(LccDetailed.bill_status, LccDetailed.bill_match_reason)
        .order_by(func.count().desc())
        .limit(MAX_GAP_GROUPS)
    )).all()

    return [{
        "status": s,
        "reason": reason,
        "count": n,
        # Dedup then cap: one passenger can hold several rows in the same gap.
        "sample_passengers": list(dict.fromkeys(samples or []))[:5],
    } for s, reason, n, samples in rows]


class BillingPartyPayload(BaseModel):
    customer_type: str | None = None      # corporate | direct
    customer_id: int | None = None
    corporate_id: int | None = None


@router.patch("/batches/{batch_id}/billing-default")
async def set_billing_default(
    batch_id: str,
    payload: BillingPartyPayload,
    apply_to: str = Query("unresolved", pattern="^(unresolved|all)$"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Set the batch's fallback party and stamp it onto rows that need one.

    `apply_to=unresolved` (the default) leaves rows a human or the resolver already
    settled; `all` re-stamps every billable row, overrides included.
    """
    batch = await _owned_batch(batch_id, db, current_user)
    ct, cust_id, corp_id = await _resolve_bill_party(
        db, current_user, payload.customer_type, payload.customer_id, payload.corporate_id
    )
    batch.default_customer_type = ct
    batch.default_customer_id = cust_id
    batch.default_corporate_id = corp_id

    rows_updated = 0
    if ct:
        conds = [LccDetailed.batch_id == batch_id, *_scope(LccDetailed, current_user),
                 LccDetailed.bill_kind != "payment"]
        if apply_to == "unresolved":
            conds.append(LccDetailed.bill_status.notin_(
                (cres.RESOLVED, cres.DEFAULTED, cres.OVERRIDDEN)
            ))
        result = await db.execute(
            update(LccDetailed).where(*conds).values(
                bill_status=cres.DEFAULTED,
                bill_customer_type=ct,
                bill_customer_id=cust_id,
                bill_corporate_id=corp_id,
                bill_match_reason="Billed to this upload's default party.",
                resolved_at=datetime.utcnow(),
                resolved_by_id=current_user.id,
            )
        )
        rows_updated = result.rowcount or 0

    if batch.resolution_status == "none":
        batch.resolution_status = "resolved"
    await _recount(db, batch)
    await db.commit()

    return {"batch_id": batch_id, "customer_type": ct, "customer_id": cust_id,
            "corporate_id": corp_id, "rows_updated": rows_updated,
            "resolved_rows": batch.resolved_rows, "unresolved_rows": batch.unresolved_rows}


@router.patch("/rows/{row_id}/billing-party")
async def set_row_billing_party(
    row_id: int,
    payload: BillingPartyPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """A human's pick for one row. Survives a re-resolve unless reset explicitly."""
    row = await db.scalar(
        select(LccDetailed).where(LccDetailed.id == row_id, *_scope(LccDetailed, current_user))
    )
    if not row:
        raise HTTPException(status_code=404, detail="Row not found.")
    if row.bill_kind == "payment":
        raise HTTPException(
            status_code=409,
            detail="This row is a payment movement — it carries no fare, so there is nothing to bill.",
        )

    ct, cust_id, corp_id = await _resolve_bill_party(
        db, current_user, payload.customer_type, payload.customer_id, payload.corporate_id
    )
    row.bill_customer_type = ct
    row.bill_customer_id = cust_id
    row.bill_corporate_id = corp_id
    row.bill_status = cres.OVERRIDDEN if ct else cres.UNRESOLVED
    row.bill_match_reason = None if ct else cres.REASON[cres.UNRESOLVED]
    row.resolved_at = datetime.utcnow()
    row.resolved_by_id = current_user.id

    batch = await _owned_batch(row.batch_id, db, current_user)
    await _recount(db, batch)
    await db.commit()

    return {"id": row.id, "bill_status": row.bill_status, "customer_type": ct,
            "customer_id": cust_id, "corporate_id": corp_id}


class BulkBillingPartyPayload(BillingPartyPayload):
    row_ids: list[int]


@router.patch("/batches/{batch_id}/billing-party-bulk")
async def set_rows_billing_party(
    batch_id: str,
    payload: BulkBillingPartyPayload,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """One party for a hand-picked set of rows — the per-row picker, in bulk.

    Stamped as OVERRIDDEN, exactly like the single-row endpoint, so a human's pick
    survives the next re-match.

    Payment movements in the selection are SKIPPED and counted, not refused. The
    single-row endpoint 409s on one because there the payment row IS the request; here,
    failing forty good rows over one payment row would be the wrong trade.
    """
    batch = await _owned_batch(batch_id, db, current_user)
    ids = _checked_ids(payload.row_ids, "use the default party above")
    await _owned_row_ids(db, batch_id, current_user, ids)

    ct, cust_id, corp_id = await _resolve_bill_party(
        db, current_user, payload.customer_type, payload.customer_id, payload.corporate_id
    )
    if not ct:
        raise HTTPException(status_code=400, detail="Pick a customer or corporate.")

    result = await db.execute(
        update(LccDetailed).where(
            LccDetailed.id.in_(ids),
            LccDetailed.batch_id == batch_id,
            *_scope(LccDetailed, current_user),
            # NULL-safe: `bill_kind != 'payment'` would drop never-resolved rows, whose
            # kind is still NULL. See lcc_billing_projection's note on the same trap.
            or_(LccDetailed.bill_kind.is_(None), LccDetailed.bill_kind != "payment"),
        ).values(
            bill_status=cres.OVERRIDDEN,
            bill_customer_type=ct,
            bill_customer_id=cust_id,
            bill_corporate_id=corp_id,
            bill_match_reason=None,
            resolved_at=datetime.utcnow(),
            resolved_by_id=current_user.id,
        )
    )
    rows_updated = result.rowcount or 0

    if batch.resolution_status == "none":
        batch.resolution_status = "resolved"
    await _recount(db, batch)
    await db.commit()

    return {"batch_id": batch_id, "customer_type": ct, "customer_id": cust_id,
            "corporate_id": corp_id, "rows_updated": rows_updated,
            "skipped_payments": len(ids) - rows_updated,
            "billable_rows": batch.billable_rows,
            "resolved_rows": batch.resolved_rows,
            "unresolved_rows": batch.unresolved_rows}


async def _owned_row_ids(
    db: AsyncSession, batch_id: str, current_user: User, ids: list[int],
) -> None:
    """Every id must be a row of THIS batch, owned by this caller, or nothing runs.

    project_batch filters on batch_id alone, so this is where the tenant boundary is
    enforced for any endpoint that takes hand-picked ids. Refusing the whole call rather
    than quietly narrowing it also stops the endpoint being used to probe which ids
    exist.
    """
    found = set((await db.execute(
        select(LccDetailed.id).where(
            LccDetailed.id.in_(ids),
            LccDetailed.batch_id == batch_id,
            *_scope(LccDetailed, current_user),
        )
    )).scalars().all())
    missing = len(ids) - len(found)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"{missing} of the selected rows are not in this upload.",
        )


def _checked_ids(row_ids: list[int] | None, action: str) -> list[int] | None:
    """Normalise a hand-picked selection. None means "the whole upload"."""
    if row_ids is None:
        return None
    ids = list(dict.fromkeys(row_ids))          # dedupe, keep the caller's order
    if not ids:
        # An EXPLICIT empty list is a mistake, not "do everything" — the None/[] split
        # is what keeps a bug in the caller from silently acting on the whole upload.
        raise HTTPException(status_code=400, detail=f"Tick at least one row, or {action}.")
    if len(ids) > MAX_SEND_ROWS:
        raise HTTPException(
            status_code=400,
            detail=f"Select at most {MAX_SEND_ROWS} rows at a time, or {action}.",
        )
    return ids


class SendToBillingPayload(BaseModel):
    """No body / null → the whole upload, synced. `row_ids` → those rows, added only."""
    row_ids: list[int] | None = Field(default=None)


@router.post("/batches/{batch_id}/send-to-billing")
async def send_to_billing(
    batch_id: str,
    payload: SendToBillingPayload | None = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Project the resolved rows into `uploaded_tickets` so billing can see them.

    Idempotent: re-running syncs rather than duplicates, and any ticket already on an
    invoice is left exactly as it was.

    With `row_ids`, only those rows are touched and nothing is ever removed — see
    project_batch's docstring for why a selective send is additive. Withdrawing a row
    from billing stays the whole-upload send's job.
    """
    batch = await _owned_batch(batch_id, db, current_user)
    if batch.status != "completed":
        raise HTTPException(status_code=409, detail="This upload is still importing.")
    if batch.resolution_status == "none":
        raise HTTPException(
            status_code=409,
            detail="Resolve this upload's customers first — nothing here has a party to bill yet.",
        )

    ids = _checked_ids(payload.row_ids if payload else None, "send the whole upload")
    if ids:
        await _owned_row_ids(db, batch_id, current_user, ids)

    result = await proj.project_batch(db, batch, row_ids=ids)
    await _recount(db, batch)
    await db.commit()
    return {"batch_id": batch_id, **result}


@router.get("/records")
async def list_records(
    request: Request,
    batch_id: str | None = Query(None),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Paginated typed rows + folded Taxes/Segments/SSR display columns.

    Narrowed by the `f.<field>` filters declared in `spec.FILTERS`; `total` and `summary`
    both cover the whole filtered set, so they stay consistent with what the grid shows.
    """
    conds = _record_conds(current_user, batch_id, request)

    # Count, pax and the column totals in one round trip — they all share the same WHERE,
    # and the summary covers the ENTIRE filtered set, not the visible page. No numeric-cast
    # guard is needed here (unlike _num() in the JSONB-backed statements router): these are
    # real NUMERIC(14,2) columns, so SUM() is exact.
    agg = (await db.execute(
        select(
            func.count().label("n"),
            func.coalesce(func.sum(LccDetailed.pax_count), 0).label("pax"),
            *[func.sum(_SUMMARY_COLS[f["field"]]).label(f"s_{f['field']}")
              for f in spec.SUMMARY_FIELDS],
        ).select_from(LccDetailed).where(*conds)
    )).one()
    total = agg.n or 0

    q = (select(LccDetailed).where(*conds)
         .order_by(LccDetailed.id.desc()).limit(limit).offset(offset))
    rows = (await db.execute(q)).scalars().all()

    fmt = ""
    if batch_id:
        b = await db.scalar(
            select(LccDetailedBatch.source_format).where(
                LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
            )
        )
        fmt = b or ""

    columns = [{"header": c["header"], "field": c["field"]} for c in spec.CORE_COLUMNS]
    columns += [
        # Declared at upload, not present in the source file — see models/tenant_airline.py.
        {"header": "Airline", "field": "__airline__"},
        {"header": "Departure Date", "field": "departure_date"},
        {"header": "Taxes Total", "field": "taxes_total"},
        {"header": "Taxes", "field": "__taxes__"},
        {"header": "Segments", "field": "__segments__"},
        {"header": "SSR", "field": "__ssr__"},
        {"header": "Format", "field": "__format__"},
    ]

    out_rows = []
    for r in rows:
        d = {c["field"]: _disp(getattr(r, c["field"], None)) for c in spec.CORE_COLUMNS}
        d["__airline__"] = " ".join(x for x in (r.airline_code, r.airline_name) if x)
        d["departure_date"] = _disp(r.departure_date)
        d["taxes_total"] = _disp(r.taxes_total)
        d["__taxes__"] = " · ".join(
            f"{t.get('code')} {t.get('amount') or ''}".strip()
            for t in (r.taxes or []) if t.get("code")
        )
        d["__segments__"] = " · ".join(_seg_str(s) for s in (r.segments or []))
        d["__ssr__"] = " · ".join(
            f"{s.get('code')} {s.get('amount') or ''}".strip()
            for s in (r.ssr or []) if s.get("code")
        )
        d["__format__"] = fmt
        d["id"] = r.id
        out_rows.append(d)

    return {
        "total": total, "limit": limit, "offset": offset,
        "columns": columns, "rows": out_rows,
        "filters": spec.FILTERS,
        "summary": {
            "fields": spec.SUMMARY_FIELDS,
            # Serialized as strings — a float round-trip would lose paise.
            "computed": {
                f["field"]: _money_str(getattr(agg, f"s_{f['field']}")) or "0"
                for f in spec.SUMMARY_FIELDS
            },
            "row_count": total,
            "pax_count": int(agg.pax or 0),
        },
    }


@router.get("/records/facets")
async def record_facets(
    batch_id: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Distinct values for every facet-backed `select` filter, for the dropdowns.

    Scoped to the batch — the question is "what payment methods are in THIS file" — but
    deliberately NOT narrowed by the other active filters, so the options don't vanish
    from under the user as they narrow. One query rather than one per column: the batch
    is scanned once and every facet aggregated in the same pass.

    Filters that declare static `options` (the `international` boolean) are absent; the
    client uses their declared options instead.
    """
    selects = [f for f in spec.FILTERS if f["type"] == "select" and not f.get("options")]
    if not selects:
        return {}

    conds = [*_scope(LccDetailed, current_user)]
    if batch_id:
        conds.append(LccDetailed.batch_id == batch_id)

    row = (await db.execute(
        select(*[func.array_agg(func.distinct(_FILTER_COLS[f["field"]])) for f in selects])
        .where(*conds)
    )).one()

    out: dict[str, list[str]] = {}
    for i, f in enumerate(selects):
        # array_agg(DISTINCT x) keeps NULL as an element, and an empty batch aggregates
        # to NULL rather than to an empty array.
        values = sorted({v for v in (row[i] or []) if v not in (None, "")})
        out[f["field"]] = values[:_MAX_FACET_VALUES]
    return out


_PREVIEW_ROWS = 2000   # rows converted to xlsx for the Excel Online preview


def _make_preview_xlsx(content: bytes, filename: str, header_row: int, nrows: int) -> bytes:
    """Read the first `nrows` rows and write them to an .xlsx (values kept as text so
    long numbers/ids/phones aren't mangled into scientific notation). Runs in a thread."""
    from openpyxl import Workbook
    df = _read_df(content, filename, header_row, nrows=nrows)
    df.dropna(how="all", inplace=True)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Sheet1")
    ws.append([str(c) for c in df.columns])
    for tup in df.itertuples(index=False, name=None):
        ws.append(["" if _clean(v) is None else _clean(v) for v in tup])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@router.get("/batches/{batch_id}/viewer-url")
async def viewer_url(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Signed URL to an .xlsx the frontend can open in the Microsoft Office web viewer
    (renders as a real spreadsheet). xlsx/xls uploads are served directly; a .csv is
    converted to an xlsx (first N rows) and cached in GCS so repeat previews are instant."""
    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch or not batch.file_url:
        raise HTTPException(status_code=404, detail="No file stored for this upload.")

    src = (batch.source_file or "").lower()
    if src.endswith((".xlsx", ".xls")):
        url = await gcs.generate_signed_url(batch.file_url, _bucket(), inline=True)
        return {"url": url, "converted": False}

    # CSV → cache a first-N-rows xlsx conversion next to the original.
    xlsx_blob = batch.file_url.rsplit("/", 1)[0] + "/_preview.xlsx"
    if not await gcs.blob_exists(xlsx_blob, _bucket()):
        content = await gcs.download_bytes(batch.file_url, _bucket())
        xlsx = await asyncio.get_event_loop().run_in_executor(
            None, _make_preview_xlsx, content, batch.source_file or "", batch.header_row or 0, _PREVIEW_ROWS
        )
        await gcs.upload_bytes(
            xlsx, xlsx_blob,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", _bucket(),
        )
    url = await gcs.generate_signed_url(xlsx_blob, _bucket(), inline=True)
    return {"url": url, "converted": True, "preview_rows": _PREVIEW_ROWS}


@router.get("/batches/{batch_id}/file-url")
async def get_batch_file_url(
    batch_id: str,
    inline: bool = Query(True, description="inline (preview) vs attachment (download)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch or not batch.file_url:
        raise HTTPException(status_code=404, detail="No file stored for this upload.")
    # For inline PREVIEW of a .csv, override the served content-type to text/plain so the
    # browser displays it instead of downloading it (browsers download text/csv even inline).
    content_type = None
    if inline and (batch.source_file or "").lower().endswith(".csv"):
        content_type = "text/plain; charset=utf-8"
    url = await gcs.generate_signed_url(
        batch.file_url, _bucket(), expiry_minutes=60, inline=inline, content_type=content_type
    )
    return {"url": url, "file_name": batch.source_file}


@router.delete("/batches/{batch_id}")
async def delete_batch(
    batch_id: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete an upload (rows cascade via FK); best-effort remove the stored file.

    Projected tickets do NOT cascade — they live in `uploaded_tickets` under their own
    statement — so they are removed here explicitly, and refused outright once any of
    them is on an invoice.
    """
    from app.models.ticket_statement import TicketStatement
    from app.models.uploaded_ticket import UploadedTicket

    batch = await db.scalar(
        select(LccDetailedBatch).where(
            LccDetailedBatch.batch_id == batch_id, *_scope(LccDetailedBatch, current_user)
        )
    )
    if not batch:
        raise HTTPException(status_code=404, detail="Upload not found.")

    if batch.billing_batch_id:
        billed = (await db.execute(
            select(func.count(), func.min(UploadedTicket.billing_id))
            .where(UploadedTicket.batch_id == batch.billing_batch_id,
                   UploadedTicket.billing_id.isnot(None))
        )).one()
        if billed[0]:
            raise HTTPException(
                status_code=409,
                detail=(f"{billed[0]} row(s) from this upload are on billing #{billed[1]}. "
                        "Delete that billing first, then delete this upload."),
            )
        await db.execute(
            delete(UploadedTicket).where(UploadedTicket.batch_id == batch.billing_batch_id)
        )
        await db.execute(
            delete(TicketStatement).where(TicketStatement.batch_id == batch.billing_batch_id)
        )

    file_url = batch.file_url
    await db.delete(batch)
    await db.commit()
    if file_url:
        for blob in (file_url, file_url.rsplit("/", 1)[0] + "/_preview.xlsx"):
            try:
                await gcs.delete_blob(blob, _bucket())
            except Exception:  # noqa: BLE001
                pass
    return {"deleted": True, "batch_id": batch_id}


@router.delete("/records/{record_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_record(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    obj = await db.scalar(
        select(LccDetailed).where(LccDetailed.id == record_id, *_scope(LccDetailed, current_user))
    )
    if not obj:
        raise HTTPException(status_code=404, detail="Record not found.")
    await db.delete(obj)
    await db.commit()
